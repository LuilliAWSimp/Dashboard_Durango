from __future__ import annotations

import unittest
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook

from app.services.water_daily_report_service import (
    SUMMARY_NOTE,
    _flow_history_drawing,
    build_daily_water_report_excel,
    build_daily_water_report_pdf,
    get_daily_water_report,
)


def period_item(
    name: str,
    *,
    operational_key: str | None = None,
    validated: float | None,
    reliable: bool,
    discarded: float = 0.0,
    discontinuity: bool = False,
    activity: str = 'Con actividad en el periodo',
) -> dict:
    return {
        'name': name,
        'operational_key': operational_key,
        'current_flow': 12.5,
        'flow_unit': 'L/s',
        'period_open_m3': 100.0,
        'period_close_m3': 200.0,
        'period_m3': validated,
        'period_m3_reliable': reliable,
        'validated_volume_m3': validated,
        'discarded_volume_m3': discarded,
        'discarded_totalizer_events': 1 if discontinuity else 0,
        'has_discontinuities': discontinuity,
        'activity': activity,
        'communication': 'Actualizado',
        'last_update': '2026-08-04T13:10:00',
        'data_status': 'invalid_totalizer' if discontinuity else 'operational',
        'samples': 60,
    }


class DurangoReportValidatedSummaryTests(unittest.TestCase):
    def setUp(self) -> None:
        self.period = {
            'wells': [
                period_item('Pozo 1', validated=0.67, reliable=False, discarded=11093.519532, discontinuity=True, activity='Con actividad'),
                period_item('Pozo 2', validated=174.98, reliable=False, discarded=4357.320313, discontinuity=True, activity='Con actividad'),
            ],
            'lines': [
                period_item('Línea 1', validated=24.21, reliable=False, discarded=500.0, discontinuity=True, activity='Con actividad'),
            ],
            'flows': [
                period_item('Lavadora Línea 2', operational_key='lavadora_linea_2', validated=0.0, reliable=True, activity='Sin actividad en el periodo'),
                period_item('Lavadora Vidrio', operational_key='lavadora_vidrio', validated=None, reliable=False, activity='Sin registros guardados'),
                period_item('Lavadora Ref Pet', operational_key='lavadora_ref_pet', validated=None, reliable=False, activity='Sin registros guardados'),
                period_item('Jarabes', operational_key='jarabes', validated=11.43, reliable=True, activity='Con actividad'),
            ],
            'summary': {
                'wells': {'total_m3': None, 'active_count': 0, 'review_count': 2},
                'lines': {'total_m3': 0.0, 'active_count': 0, 'review_count': 1},
                'flows': {'total_m3': 0.0, 'active_count': 0, 'review_count': 0},
            },
            'source_status': 'readings_minute',
        }

    def review_payload(self, *, shifts: list | None = None) -> dict:
        return {
            'source_status': self.period.get('source_status'),
            'validated_segment_start': '2026-08-04T18:16:00',
            'crosses_scada_cutover': False,
            'legacy_notice': None,
            'modules': {
                'wells': {'items': self.period['wells'], 'summary': self.period['summary']['wells']},
                'lines': {'items': self.period['lines'], 'summary': self.period['summary']['lines']},
                'flows': {'items': self.period['flows'], 'summary': self.period['summary']['flows']},
            },
            'shifts': {'shifts': shifts or []},
        }

    def build_report(self) -> dict:
        with patch('app.services.water_daily_report_service.get_daily_water_review', return_value=self.review_payload()), patch(
            'app.services.water_daily_report_service.get_water_history_module',
            side_effect=lambda module, start_date, end_date, aggregation: {
                'module': module,
                'start_date': start_date,
                'end_date': end_date,
                'aggregation': aggregation,
                'series': [],
                'source_status': 'no_data',
                'has_future_intervals': False,
            },
        ):
            return get_daily_water_report('2026-08-04')

    def test_summary_includes_partial_validated_volume_but_not_discarded_events(self) -> None:
        report = self.build_report()
        summary = report['summary']
        self.assertAlmostEqual(summary['well_validated_volume_m3'], 175.65, places=6)
        self.assertAlmostEqual(summary['line_validated_volume_m3'], 24.21, places=6)
        self.assertAlmostEqual(summary['flow_validated_volume_m3'], 11.43, places=6)
        self.assertEqual(summary['washer_validated_volume_m3'], 0.0)
        self.assertAlmostEqual(summary['jarabes_validated_volume_m3'], 11.43, places=6)
        self.assertAlmostEqual(summary['total_validated_operational_m3'], 211.29, places=6)
        self.assertAlmostEqual(summary['discarded_volume_m3'], 15950.839845, places=6)
        self.assertEqual(summary['review_count'], 0)
        self.assertEqual(summary['partial_validation_count'], 0)
        self.assertEqual(summary['validated_items_count'], 5)
        self.assertEqual(summary['note'], SUMMARY_NOTE)
        self.assertEqual(report['notes'], [])
        self.assertNotIn('Lavadora Línea 2', [item['name'] for item in report['production_lines']['rows']])
        self.assertEqual([item['name'] for item in report['operational_flows']['rows']].count('Lavadora Línea 2'), 1)
        self.assertEqual([item['name'] for item in report['washers']['rows']], ['Lavadora Línea 2', 'Lavadora Vidrio', 'Lavadora Ref Pet'])
        self.assertEqual([item['name'] for item in report['jarabes']['rows']], ['Jarabes'])

    def test_excel_uses_numeric_validated_values_and_same_summary(self) -> None:
        report = self.build_report()
        content, _ = build_daily_water_report_excel(report)
        workbook = load_workbook(BytesIO(content), data_only=False)
        summary_sheet = workbook['Resumen']
        summary_values = {summary_sheet.cell(row, 1).value: summary_sheet.cell(row, 2).value for row in range(2, summary_sheet.max_row + 1)}
        self.assertAlmostEqual(summary_values['Volumen validado de pozos (m³)'], 175.65, places=6)
        self.assertAlmostEqual(summary_values['Volumen validado de líneas (m³)'], 24.21, places=6)
        self.assertEqual(summary_values['Volumen validado de lavadoras (m³)'], 0)
        self.assertAlmostEqual(summary_values['Volumen validado de Jarabes (m³)'], 11.43, places=6)
        self.assertAlmostEqual(summary_values['Subtotal validado operativo (m³)'], 211.29, places=6)
        wells_sheet = workbook['Pozos']
        self.assertIsInstance(wells_sheet['E2'].value, (int, float))
        self.assertAlmostEqual(wells_sheet['E2'].value, 0.67, places=6)
        self.assertEqual(wells_sheet['F2'].value, 'Validado')
        self.assertEqual(wells_sheet['G2'].value, 'Con actividad')
        self.assertEqual(workbook.sheetnames[:6], ['Resumen', 'Pozos', 'Líneas', 'Lavadoras', 'Jarabes', 'Turnos'])

    def test_pdf_is_generated_from_same_report_object(self) -> None:
        report = self.build_report()
        content, filename = build_daily_water_report_pdf(report)
        self.assertTrue(content.startswith(b'%PDF'))
        self.assertEqual(filename, 'reporte-diario-control-hidrico-durango-2026-08-04.pdf')

    def test_preview_skips_histories_and_shifts(self) -> None:
        with patch('app.services.water_daily_report_service.get_daily_water_review', return_value=self.review_payload()) as review, patch(
            'app.services.water_daily_report_service.get_water_history_module'
        ) as history:
            report = get_daily_water_report('2026-08-04', include_history=False, include_shifts=False)
        history.assert_not_called()
        self.assertFalse(review.call_args.kwargs['include_shifts'])
        self.assertFalse(report['includes_history'])
        self.assertFalse(report['includes_shifts'])
        self.assertEqual(report['report_source'], 'daily_review')
        self.assertEqual(report['shifts'], [])
        self.assertEqual(report['history']['wells'], {})

    def test_full_report_obtains_all_histories_and_reuses_daily_review_shifts(self) -> None:
        shift = {'id': 'shift_1', 'name': 'T1', 'schedule': '00:00-07:00'}
        with patch('app.services.water_daily_report_service.get_daily_water_review', return_value=self.review_payload(shifts=[shift])) as review, patch(
            'app.services.water_daily_report_service.get_water_history_module',
            return_value={'series': [], 'aggregation': 'quarter_hour'},
        ) as history:
            report = get_daily_water_report('2026-08-04')
        self.assertTrue(review.call_args.kwargs['include_shifts'])
        self.assertEqual(history.call_count, 3)
        self.assertTrue(report['includes_history'])
        self.assertTrue(report['includes_shifts'])
        self.assertEqual(report['shifts'], [shift])

    def test_pdf_flow_axis_unit_is_vertical_and_separated_from_ticks(self) -> None:
        drawing = _flow_history_drawing(
            {
                'aggregation': 'quarter_hour',
                'series': [{
                    'name': 'Pozo 1',
                    'points': [
                        {'bucket_start': '2026-08-04T00:00:00', 'samples': 15, 'flow_avg_lps': 0.0},
                        {'bucket_start': '2026-08-04T00:15:00', 'samples': 15, 'flow_avg_lps': 224.9},
                    ],
                }],
            },
            width=527,
            height=181,
            single_day=True,
        )
        strings = [item for item in drawing.contents if item.__class__.__name__ == 'String']
        axis = next(item for item in strings if item.text == 'L/s')
        numeric_ticks = [item for item in strings if item.text in {'0.0', '56.2', '112.4', '168.7', '251.9'}]
        self.assertEqual(axis.angle, 90)
        self.assertTrue(numeric_ticks)
        self.assertLess(axis.x, min(item.x for item in numeric_ticks) - 10)


if __name__ == '__main__':
    unittest.main()
