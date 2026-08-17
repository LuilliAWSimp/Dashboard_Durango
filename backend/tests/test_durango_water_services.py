from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook

from app.services.durango_capabilities import FLOWS, JARABES, LAVADORAS, LINE_FLOWS, LINES, WELLS
from app.services.totalizer_quality import analyze_totalizer_series
from app.services.water_history_service import _build_points
from app.services.water_shift_service import SHIFT_DEFINITIONS, _window
from app.services.water_period_service import build_period_item
import app.services.water_daily_report_service as report_service


def test_confirmed_mappings_keep_durango_order():
    assert [(item['sensor_id'], item['name']) for item in WELLS] == [(1001, 'Pozo 1'), (1051, 'Pozo 2')]
    assert [(item['sensor_id'], item['name']) for item in LINES] == [
        (2002, 'Línea 1'), (2006, 'Línea 3'), (2008, 'Línea 4'), (2010, 'Línea 5')
    ]
    assert [(item['sensor_id'], item['name']) for item in LINE_FLOWS] == [(2004, 'Lavadora Línea 2')]
    assert [(item['operational_key'], item['name']) for item in LAVADORAS] == [
        ('lavadora_vidrio', 'Lavadora Vidrio'), ('lavadora_ref_pet', 'Lavadora Ref Pet')
    ]
    assert [(item['sensor_id'], item['name']) for item in JARABES] == [(3004, 'Jarabes')]


def test_totalizer_ignores_intermittent_zero_without_double_counting():
    result = analyze_totalizer_series([
        ('2026-08-01T00:00:00', 100.0),
        ('2026-08-01T00:01:00', 0.0),
        ('2026-08-01T00:02:00', 102.0),
        ('2026-08-01T00:03:00', 105.0),
    ], sensor_id=2002)
    assert result.reliable is True
    assert result.volume_m3 == 5.0


def test_totalizer_restart_is_review_not_negative_volume():
    result = analyze_totalizer_series([
        ('2026-08-01T00:00:00', 100.0),
        ('2026-08-01T01:00:00', 110.0),
        ('2026-08-01T02:00:00', 3.0),
    ], sensor_id=2002)
    assert result.reliable is False
    assert result.volume_m3 == 10.0
    assert result.validated_volume_m3 == 10.0
    assert result.has_discontinuities is True
    assert result.status == 'invalid_totalizer'


def test_empty_history_preserves_nulls_and_does_not_invent_current_point():
    points = _build_points(
        2002,
        'hourly',
        datetime(2026, 8, 5, 0, 0),
        datetime(2026, 8, 5, 3, 0),
        [],
    )
    assert len(points) == 3
    assert all(point['samples'] == 0 for point in points)
    assert all(point['flow_avg_lps'] is None for point in points)
    assert all(point['volume_m3'] is None for point in points)
    assert all(point['data_status'] == 'no_data' for point in points)


def test_shift_windows_are_non_overlapping():
    day = date(2026, 8, 1)
    windows = [_window(day, definition) for definition in SHIFT_DEFINITIONS]
    assert windows[0] == (datetime(2026, 8, 1, 0, 0), datetime(2026, 8, 1, 7, 0))
    assert windows[1] == (datetime(2026, 8, 1, 7, 0), datetime(2026, 8, 1, 15, 0))
    assert windows[2] == (datetime(2026, 8, 1, 15, 0), datetime(2026, 8, 2, 0, 0))
    assert windows[0][1] == windows[1][0]
    assert windows[1][1] == windows[2][0]


def _period_fixture():
    contracts = [*WELLS, *LINES, *FLOWS]
    rows = []
    for index, contract in enumerate(contracts, start=1):
        rows.append({
            'sensor_id': contract.get('sensor_id'),
            'operational_key': contract['operational_key'],
            'name': contract['name'],
            'nombre': contract['name'],
            'module': contract['group'],
            'current_flow': float(index),
            'flow_unit': contract['flow_unit'],
            'period_open_m3': 1000.0 + index,
            'period_close_m3': 1001.0 + index,
            'current_totalizer_m3': 1001.0 + index,
            'period_m3': 1.0,
            'period_m3_reliable': True,
            'activity': 'Con actividad en el periodo',
            'communication': 'Actualizado',
            'last_update': '2026-08-01T12:00:00',
            'data_status': 'operational',
            'samples': 60,
        })
    groups = {key: [row for row in rows if row['module'] == key] for key in ('well', 'line', 'flow')}

    def summary(group_rows):
        return {
            'total_m3': float(len(group_rows)),
            'active_count': len(group_rows),
            'inactive_count': 0,
            'review_count': 0,
            'coverage_available': len(group_rows),
            'coverage_total': len(group_rows),
        }

    return {
        'wells': groups['well'],
        'lines': groups['line'],
        'flows': groups['flow'],
        'summary': {key + 's' if key != 'flow' else 'flows': summary(groups[key]) for key in ()},
        'source_status': 'operational',
    } | {'summary': {'wells': summary(groups['well']), 'lines': summary(groups['line']), 'flows': summary(groups['flow'])}}


def _history_fixture(module, start_date, end_date, aggregation, **_):
    contracts = {'well': WELLS, 'line': LINES, 'flow': FLOWS}[module]
    return {
        'module': module,
        'start_date': start_date,
        'end_date': end_date,
        'aggregation': aggregation,
        'source_status': 'operational',
        'has_future_intervals': False,
        'series': [
            {
                'sensor_id': item.get('sensor_id'),
                'operational_key': item['operational_key'],
                'name': item['name'],
                'has_data': True,
                'points': [{
                    'bucket_start': f'{start_date}T00:00:00',
                    'bucket_end': f'{start_date}T00:15:00',
                    'aggregation': aggregation,
                    'samples': 15,
                    'samples_received': 15,
                    'samples_expected': 15,
                    'coverage_percent': 100.0,
                    'active_minutes': 0,
                    'flow_avg_lps': 0.0,
                    'flow_active_avg_lps': None,
                    'flow_min_lps': 0.0,
                    'flow_max_lps': 0.0,
                    'totalizer_open_m3': 100.0,
                    'totalizer_close_m3': 100.0,
                    'validated_volume_m3': 0.0,
                    'interval_state': 'Apagado con datos',
                    'data_status': 'operational',
                }],
            }
            for item in contracts
        ],
    }


def test_pdf_excel_and_email_attachment_use_same_report_structure(monkeypatch):
    fixture = _period_fixture()
    monkeypatch.setattr(report_service, 'get_period_data', lambda *args, **kwargs: fixture)
    monkeypatch.setattr(report_service, 'get_shift_consumption_data', lambda *args, **kwargs: {'shifts': []})
    monkeypatch.setattr(report_service, 'get_water_history_module', _history_fixture)
    report = report_service.get_daily_water_report('2026-08-01')
    pdf_bytes, pdf_name = report_service.build_daily_water_report_pdf(report)
    excel_bytes, excel_name = report_service.build_daily_water_report_excel(report)
    assert pdf_bytes.startswith(b'%PDF')
    assert pdf_name == 'reporte-diario-control-hidrico-durango-2026-08-01.pdf'
    assert excel_name == 'reporte-diario-control-hidrico-durango-2026-08-01.xlsx'
    workbook = load_workbook(BytesIO(excel_bytes), data_only=True)
    assert workbook.sheetnames[:8] == ['Resumen', 'Pozos', 'Líneas', 'Flujos', 'Turnos', 'Histórico Pozos', 'Histórico Líneas', 'Histórico Flujos']
    assert workbook['Pozos'].max_row == 3
    assert workbook['Líneas'].max_row == 5
    assert workbook['Flujos'].max_row == 5
    assert workbook['Histórico Pozos']['E2'].value == 0
    assert workbook['Histórico Pozos']['J2'].value == 15
    assert report['history']['aggregation'] == 'quarter_hour'
    assert [row['name'] for row in report['production_lines']['rows']] == ['Línea 1', 'Línea 3', 'Línea 4', 'Línea 5']
    assert [row['name'] for row in report['operational_flows']['rows']] == ['Lavadora Línea 2', 'Lavadora Vidrio', 'Lavadora Ref Pet', 'Jarabes']


def test_shift_boundary_minutes_belong_to_expected_turn():
    day = date(2026, 8, 1)
    windows = [(*_window(day, definition), definition['id']) for definition in SHIFT_DEFINITIONS]

    def assigned(value: datetime) -> str | None:
        return next((shift_id for start, end, shift_id in windows if start <= value < end), None)

    assert assigned(datetime(2026, 8, 1, 6, 59)) == 'shift_1'
    assert assigned(datetime(2026, 8, 1, 7, 0)) == 'shift_2'
    assert assigned(datetime(2026, 8, 1, 14, 59)) == 'shift_2'
    assert assigned(datetime(2026, 8, 1, 15, 0)) == 'shift_3'
    next_day_windows = [(*_window(day + timedelta(days=1), definition), definition['id']) for definition in SHIFT_DEFINITIONS]
    assert next((shift_id for start, end, shift_id in next_day_windows if start <= datetime(2026, 8, 2, 0, 0) < end), None) == 'shift_1'


def test_activity_uses_totalizer_movement_even_when_latest_flow_is_zero():
    contract = {'sensor_id': 2002, 'display_name': 'Línea 1', 'group': 'line', 'flow_unit': 'L/s'}
    rows = [
        {'operational_ts': datetime(2026, 8, 1, 0, 0), 'instant_value': 2.0, 'total_value': 100.0},
        {'operational_ts': datetime(2026, 8, 1, 1, 0), 'instant_value': 0.0, 'total_value': 105.0},
    ]
    item = build_period_item(contract, rows, None, date(2026, 8, 1))
    assert item['current_flow'] == 0.0
    assert item['period_m3'] == 5.0
    assert item['activity'] == 'Con actividad'


def test_recent_samples_without_totalizer_movement_are_inactive_not_offline():
    contract = {'sensor_id': 2002, 'display_name': 'Línea 1', 'group': 'line', 'flow_unit': 'L/s'}
    stamp = datetime.now().replace(second=0, microsecond=0)
    rows = [
        {'operational_ts': stamp - timedelta(minutes=1), 'instant_value': 0.0, 'total_value': 100.0},
        {'operational_ts': stamp, 'instant_value': 0.0, 'total_value': 100.0},
    ]
    item = build_period_item(contract, rows, None, datetime.now().date())
    assert item['period_m3'] == 0.0
    assert item['activity'] == 'Sin actividad'
    assert item['communication'] == 'Actualizado'
