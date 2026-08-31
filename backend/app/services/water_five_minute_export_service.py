from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from io import BytesIO
import re
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import SQLAlchemyError

from app.services.durango_capabilities import (
    DURANGO_SCADA_CUTOVER_LOCAL,
    LOCAL_TIMEZONE,
    LAVADORAS,
    SENSORS_BY_MODULE,
    clamp_to_validated_segment,
    identity_key,
    is_jarabes_identity,
    item_contract,
)
from app.services.durango_jarabes_service import query_jarabes_previous_reading, query_jarabes_rows
from app.services.durango_lavadoras_service import (
    build_lavadora_period_item,
    query_lavadora_previous_readings,
    query_lavadora_rows,
)
from app.services.plant_time import local_now_naive
from app.services.water_period_service import build_period_item, query_previous_closes, query_readings_window

Module = Literal['well', 'line', 'flow']
MAX_EXPORT_DAYS = 3
BUCKET_MINUTES = 5
LAVADORA_KEYS = {str(item['operational_key']) for item in LAVADORAS}


class FiveMinuteExportError(RuntimeError):
    def __init__(self, message: str, *, status: str = 'sql_error') -> None:
        super().__init__(message)
        self.status = status


@dataclass(frozen=True)
class ExportRange:
    start_day: date
    end_day: date
    requested_start: datetime
    requested_end: datetime
    local_start: datetime
    local_end: datetime
    crosses_scada_cutover: bool

    @property
    def inclusive_days(self) -> int:
        return (self.end_day - self.start_day).days + 1


def _parse_date(value: Any, label: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(str(value)[:10]).date()
    except (TypeError, ValueError) as exc:
        raise ValueError(f'La fecha {label} no es valida.') from exc


def _floor_five_minutes(value: datetime) -> datetime:
    minute = (value.minute // BUCKET_MINUTES) * BUCKET_MINUTES
    return value.replace(minute=minute, second=0, microsecond=0)


def _next_five_minute_boundary(value: datetime) -> datetime:
    floored = _floor_five_minutes(value)
    return floored if floored == value else floored + timedelta(minutes=BUCKET_MINUTES)


def _export_range(start_date: Any, end_date: Any, *, now: datetime | None = None) -> ExportRange:
    start_day = _parse_date(start_date, 'inicial')
    end_day = _parse_date(end_date, 'final')
    if start_day > end_day:
        raise ValueError('La fecha inicial no puede ser posterior a la fecha final.')
    inclusive_days = (end_day - start_day).days + 1
    if inclusive_days > MAX_EXPORT_DAYS:
        raise ValueError('La exportacion de 5 minutos permite un maximo de 3 dias calendario.')

    requested_start = datetime.combine(start_day, time.min)
    requested_end = datetime.combine(end_day + timedelta(days=1), time.min)
    now_local = (now or local_now_naive()).replace(tzinfo=None)
    completed_end = min(requested_end, _floor_five_minutes(now_local)) if end_day >= now_local.date() else requested_end
    validated_start, validated_end, legacy_only, crosses = clamp_to_validated_segment(requested_start, completed_end)
    if legacy_only or validated_end <= validated_start:
        raise ValueError('El rango seleccionado no contiene intervalos del segmento validado posterior al corte SCADA.')
    if validated_end <= validated_start:
        raise ValueError('El periodo seleccionado todavia no contiene intervalos completos de 5 minutos.')
    return ExportRange(
        start_day=start_day,
        end_day=end_day,
        requested_start=requested_start,
        requested_end=requested_end,
        local_start=validated_start,
        local_end=validated_end,
        crosses_scada_cutover=crosses,
    )


def _module_contract(module: str, element_id: Any) -> tuple[Module, dict[str, Any], str]:
    normalized_module = str(module or '').strip().lower()
    if normalized_module not in SENSORS_BY_MODULE:
        raise ValueError('El modulo de exportacion debe ser well, line o flow.')
    identity = identity_key(element_id)
    allowed = {identity_key(value) for value in SENSORS_BY_MODULE[normalized_module]}
    if identity not in allowed:
        raise ValueError('El elemento solicitado no pertenece al contrato operativo de Durango.')
    contract = item_contract(element_id)
    if not contract.get('enabled', True):
        raise ValueError('El elemento solicitado no esta habilitado en el contrato operativo de Durango.')
    return normalized_module, contract, identity  # type: ignore[return-value]


def _num(value: Any) -> float | None:
    if value in (None, ''):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if parsed == parsed else None


def _stamp(row: dict[str, Any] | None) -> datetime | None:
    value = (row or {}).get('operational_ts')
    return value if isinstance(value, datetime) else None


def _last_positive_row(rows: list[dict[str, Any]], fallback: dict[str, Any] | None = None) -> dict[str, Any] | None:
    previous = dict(fallback) if fallback else None
    for row in rows:
        total = _num(row.get('total_value'))
        if total is not None and total > 0 and _stamp(row) is not None:
            previous = dict(row)
    return previous


def _sensor_previous_tuple(row: dict[str, Any] | None) -> tuple[datetime | None, float | None, float | None] | None:
    if not row:
        return None
    return (_stamp(row), _num(row.get('total_value')), _num(row.get('instant_value')))


def _load_source_rows(contract: dict[str, Any], export_range: ExportRange) -> tuple[list[dict[str, Any]], dict[str, Any] | None, str]:
    key = str(contract.get('operational_key') or '')
    try:
        if key in LAVADORA_KEYS:
            grouped = query_lavadora_rows(export_range.local_start, export_range.local_end)
            previous = query_lavadora_previous_readings(export_range.local_start).get(key)
            return list(grouped.get(key) or []), previous, 'dbo.SensorsBOS_Lavadoras'
        if is_jarabes_identity(key):
            rows = query_jarabes_rows(export_range.local_start, export_range.local_end)
            previous = query_jarabes_previous_reading(export_range.local_start)
            return rows, previous, 'dbo.SensorsBOS_Tanque'

        sensor_id = contract.get('sensor_id')
        if sensor_id is None:
            raise ValueError('El elemento no tiene una fuente historica resoluble.')
        sensor = int(sensor_id)
        rows = query_readings_window([sensor], export_range.local_start, export_range.local_end)
        previous_tuple = query_previous_closes([sensor], export_range.local_start).get(sensor)
        previous = None
        if previous_tuple:
            previous = {
                'operational_ts': previous_tuple[0],
                'total_value': previous_tuple[1],
                'instant_value': previous_tuple[2],
                'source': 'iot.readings_minute',
                'period_source': 'iot.readings_minute',
            }
        return rows, previous, 'iot.readings_minute'
    except SQLAlchemyError as exc:
        raise FiveMinuteExportError('No fue posible consultar la fuente historica del elemento.', status='sql_error') from exc
    except Exception as exc:
        if isinstance(exc, (ValueError, FiveMinuteExportError)):
            raise
        status = getattr(exc, 'status', 'sql_error')
        raise FiveMinuteExportError(str(exc) or 'No fue posible consultar la fuente historica del elemento.', status=status) from exc


def _first_bucket_end(cursor: datetime) -> datetime:
    boundary = _next_five_minute_boundary(cursor)
    return boundary if boundary > cursor else cursor + timedelta(minutes=BUCKET_MINUTES)


def _quality_label(item: dict[str, Any]) -> str:
    return str(item.get('quality_label') or item.get('validation') or item.get('coverage_status') or 'Sin datos')


def _build_bucket_row(
    *,
    module: Module,
    contract: dict[str, Any],
    bucket_rows: list[dict[str, Any]],
    previous: dict[str, Any] | None,
    bucket_start: datetime,
    bucket_end: datetime,
    source_status: str,
) -> dict[str, Any]:
    key = str(contract.get('operational_key') or '')
    if key in LAVADORA_KEYS or is_jarabes_identity(key):
        item = build_lavadora_period_item(
            contract,
            bucket_rows,
            bucket_end.date(),
            window_start=bucket_start,
            window_end=bucket_end,
            previous_reading=previous,
        )
    else:
        item = build_period_item(
            contract,
            bucket_rows,
            _sensor_previous_tuple(previous),
            bucket_end.date(),
            window_start=bucket_start,
            window_end=bucket_end,
        )

    reconciled_volume = _num(item.get('reconciled_validated_volume_m3'))
    reliable = bool(item.get('reconciled_volume_reliable') or item.get('quality_volume_reliable'))
    return {
        'element': str(contract.get('display_name') or contract.get('name') or key),
        'element_id': contract.get('sensor_id') if contract.get('sensor_id') is not None else key,
        'sensor_id': contract.get('sensor_id'),
        'operational_key': key,
        'start_local': bucket_start,
        'end_local': bucket_end,
        'flow_unit': str(contract.get('flow_unit') or 'L/s'),
        'flow_avg': _num(item.get('flow_avg')),
        'flow_min': _num(item.get('flow_min')),
        'flow_max': _num(item.get('flow_max')),
        'totalizer_open_m3': _num(item.get('reconciled_open_m3')),
        'totalizer_close_m3': _num(item.get('reconciled_close_m3')),
        'validated_volume_m3': reconciled_volume,
        'reported_volume_m3': reconciled_volume if reliable else None,
        'volume_reliable': reliable,
        'samples': int(item.get('samples_received') or item.get('samples') or 0),
        'samples_expected': int(item.get('samples_expected') or 0),
        'coverage_pct': _num(item.get('coverage_percent')),
        'coverage_status': str(item.get('coverage_status') or ''),
        'opening_source': str(item.get('opening_source') or 'no_data'),
        'boundary_complete': bool(item.get('boundary_complete')),
        'has_discontinuities': bool(item.get('reconciled_has_discontinuities') or item.get('has_discontinuities')),
        'quality_status': str(item.get('quality_status') or ''),
        'quality_label': _quality_label(item),
        'source_status': source_status,
    }


def _build_rows(
    *,
    module: Module,
    contract: dict[str, Any],
    export_range: ExportRange,
    source_rows: list[dict[str, Any]],
    initial_previous: dict[str, Any] | None,
    source_status: str,
) -> list[dict[str, Any]]:
    rows = [dict(row) for row in source_rows if isinstance(row.get('operational_ts'), datetime)]
    rows.sort(key=lambda row: row['operational_ts'])
    result: list[dict[str, Any]] = []
    previous = dict(initial_previous) if initial_previous else None
    index = 0
    cursor = export_range.local_start

    while cursor < export_range.local_end:
        bucket_end = min(_first_bucket_end(cursor), export_range.local_end)
        if bucket_end <= cursor:
            break
        bucket_rows: list[dict[str, Any]] = []
        while index < len(rows) and rows[index]['operational_ts'] < bucket_end:
            row = rows[index]
            if row['operational_ts'] >= cursor:
                bucket_rows.append(row)
            index += 1
        result.append(_build_bucket_row(
            module=module,
            contract=contract,
            bucket_rows=bucket_rows,
            previous=previous,
            bucket_start=cursor,
            bucket_end=bucket_end,
            source_status=source_status,
        ))
        previous = _last_positive_row(bucket_rows, previous)
        cursor = bucket_end
    return result


def get_five_minute_export_data(*, module: str, element_id: Any, start_date: Any, end_date: Any) -> dict[str, Any]:
    typed_module, contract, identity = _module_contract(module, element_id)
    export_range = _export_range(start_date, end_date)
    source_rows, previous, source_status = _load_source_rows(contract, export_range)
    rows = _build_rows(
        module=typed_module,
        contract=contract,
        export_range=export_range,
        source_rows=source_rows,
        initial_previous=previous,
        source_status=source_status,
    )
    return {
        'plant': 'Planta Durango',
        'module': typed_module,
        'element_id': contract.get('sensor_id') if contract.get('sensor_id') is not None else identity,
        'sensor_id': contract.get('sensor_id'),
        'operational_key': str(contract.get('operational_key') or identity),
        'name': str(contract.get('display_name') or contract.get('name') or identity),
        'flow_unit': str(contract.get('flow_unit') or 'L/s'),
        'start_date': export_range.start_day.isoformat(),
        'end_date': export_range.end_day.isoformat(),
        'effective_start_local': export_range.local_start.isoformat(timespec='seconds'),
        'effective_end_local': export_range.local_end.isoformat(timespec='seconds'),
        'time_zone': LOCAL_TIMEZONE,
        'bucket_minutes': BUCKET_MINUTES,
        'source_status': source_status,
        'crosses_scada_cutover': export_range.crosses_scada_cutover,
        'scada_cutover_local': DURANGO_SCADA_CUTOVER_LOCAL.isoformat(timespec='seconds'),
        'rows': rows,
    }


def _safe_token(value: str) -> str:
    return re.sub(r'[^A-Za-z0-9_-]+', '_', value.strip()).strip('_') or 'elemento'


def _round_or_none(value: Any, digits: int = 4) -> float | None:
    parsed = _num(value)
    return round(parsed, digits) if parsed is not None else None


def build_five_minute_excel(payload: dict[str, Any]) -> tuple[bytes, str]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '5 minutos'

    name = str(payload.get('name') or 'Elemento')
    element_id = str(payload.get('element_id') or payload.get('operational_key') or '')
    sheet['A1'] = 'ARCA Durango - Historico conciliado cada 5 minutos'
    sheet['A1'].font = Font(bold=True, size=14)
    sheet.merge_cells('A1:P1')
    metadata = [
        ('Elemento', name),
        ('Identidad', element_id),
        ('Modulo', str(payload.get('module') or '')),
        ('Rango solicitado', f"{payload.get('start_date')} a {payload.get('end_date')}"),
        ('Ventana efectiva', f"{payload.get('effective_start_local')} a {payload.get('effective_end_local')}"),
        ('Zona horaria', str(payload.get('time_zone') or LOCAL_TIMEZONE)),
        ('Fuente', str(payload.get('source_status') or '')),
        ('Regla', '[T0,T1) con apertura previa; la apertura no cuenta como muestra'),
    ]
    for row_index, (label, value) in enumerate(metadata, start=2):
        sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row_index, column=2, value=value)

    headers = [
        'Elemento', 'Identidad', 'Inicio local', 'Fin local', 'Flujo promedio', 'Flujo minimo', 'Flujo maximo',
        'Apertura totalizador (m3)', 'Cierre totalizador (m3)', 'Volumen conciliado (m3)',
        'Volumen reportable (m3)', 'Muestras', 'Esperadas', 'Cobertura (%)', 'Calidad', 'Fuente apertura',
    ]
    header_row = 11
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='1F4E78')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    rows = list(payload.get('rows') or [])
    for row_index, item in enumerate(rows, start=header_row + 1):
        values = [
            item.get('element'), item.get('element_id'), item.get('start_local'), item.get('end_local'),
            _round_or_none(item.get('flow_avg')), _round_or_none(item.get('flow_min')), _round_or_none(item.get('flow_max')),
            _round_or_none(item.get('totalizer_open_m3')), _round_or_none(item.get('totalizer_close_m3')),
            _round_or_none(item.get('validated_volume_m3')), _round_or_none(item.get('reported_volume_m3')),
            item.get('samples'), item.get('samples_expected'), _round_or_none(item.get('coverage_pct'), 2),
            item.get('quality_label'), item.get('opening_source'),
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column, value=value)
        for column in (3, 4):
            sheet.cell(row=row_index, column=column).number_format = 'yyyy-mm-dd hh:mm'

    sheet.freeze_panes = f'A{header_row + 1}'
    sheet.auto_filter.ref = f'A{header_row}:P{header_row + max(len(rows), 1)}'
    widths = [24, 18, 19, 19, 16, 14, 14, 22, 22, 22, 22, 12, 12, 14, 24, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    reconciliation = workbook.create_sheet('Conciliacion')
    reconciliation.append(['Concepto', 'Valor'])
    reconciliation['A1'].font = reconciliation['B1'].font = Font(bold=True, color='FFFFFF')
    reconciliation['A1'].fill = reconciliation['B1'].fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    reliable_rows = [row for row in rows if row.get('volume_reliable') and _num(row.get('reported_volume_m3')) is not None]
    partial_rows = [row for row in rows if row.get('samples') and not row.get('volume_reliable')]
    no_data_rows = [row for row in rows if not row.get('samples')]
    reconciliation_rows = [
        ('Elemento', name),
        ('Intervalos generados', len(rows)),
        ('Intervalos con volumen reportable', len(reliable_rows)),
        ('Intervalos con datos pero no reportables', len(partial_rows)),
        ('Intervalos sin datos', len(no_data_rows)),
        ('Subtotal reportable (m3)', round(sum(float(row.get('reported_volume_m3') or 0.0) for row in reliable_rows), 4)),
        ('Criterio', 'No convertir huecos ni volumen no confiable a 0 m3.'),
        ('Corte SCADA', str(payload.get('scada_cutover_local') or '')),
    ]
    for item in reconciliation_rows:
        reconciliation.append(list(item))
    reconciliation.column_dimensions['A'].width = 40
    reconciliation.column_dimensions['B'].width = 72

    output = BytesIO()
    workbook.save(output)
    filename = f"ARCA_Durango_{_safe_token(name)}_5min_{payload.get('start_date')}_{payload.get('end_date')}.xlsx"
    return output.getvalue(), filename
