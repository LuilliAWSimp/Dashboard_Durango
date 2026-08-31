from __future__ import annotations

from datetime import date, datetime
from html import escape
from io import BytesIO
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing, Line, Path as DrawingPath, Rect, String
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import HRFlowable, Image, KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.services.durango_capabilities import FLOWS, JARABES, LOCAL_TIMEZONE
from app.services.water_daily_review_service import DailyReviewError, get_daily_water_review
from app.services.water_history_service import WaterHistoryError, get_water_history_module
from app.services.water_period_service import WaterPeriodError, get_period_data

LOCAL_ZONE = ZoneInfo(LOCAL_TIMEZONE)
SUMMARY_NOTE = (
    'Los volúmenes mostrados consideran únicamente incrementos validados. '
    'Los eventos descartados no se incluyen en los totales.'
)

LAVADORA_KEYS = {str(item.get('operational_key')) for item in FLOWS if str(item.get('operational_key') or '').startswith('lavadora_')}
JARABES_KEYS = {str(item.get('operational_key')) for item in JARABES}


def _operational_key(item: dict[str, Any]) -> str:
    return str(item.get('operational_key') or item.get('operationalKey') or '').strip().lower()


def _is_lavadora(item: dict[str, Any]) -> bool:
    return _operational_key(item) in LAVADORA_KEYS


def _is_jarabes(item: dict[str, Any]) -> bool:
    return _operational_key(item) in JARABES_KEYS


def _split_flow_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    lavadoras = [row for row in rows if _is_lavadora(row)]
    jarabes = [row for row in rows if _is_jarabes(row)]
    return lavadoras, jarabes


def _filter_history(history: dict[str, Any], allowed_keys: set[str]) -> dict[str, Any]:
    clone = dict(history or {})
    clone['series'] = [
        series for series in list((history or {}).get('series') or [])
        if str(series.get('operational_key') or '').strip().lower() in allowed_keys
    ]
    return clone


class ReportDataUnavailableError(RuntimeError):
    pass


def _parse_date(value: Any = None) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value:
        try:
            return datetime.fromisoformat(str(value)[:10]).date()
        except ValueError:
            pass
    return datetime.now(LOCAL_ZONE).date()


def _fmt_number(value: Any, decimals: int = 2) -> str:
    if value is None:
        return 'No disponible'
    try:
        return f'{float(value):,.{decimals}f}'
    except (TypeError, ValueError):
        return 'No disponible'


def _fmt_volume(value: Any) -> str:
    return 'No disponible' if value is None else f'{_fmt_number(value)} m³'


def _fmt_date(value: Any) -> str:
    if not value:
        return 'Sin lectura'
    try:
        parsed = datetime.fromisoformat(str(value).replace('Z', ''))
        return parsed.strftime('%d/%m/%Y %H:%M')
    except ValueError:
        return str(value)


def _as_float(value: Any) -> float | None:
    if value is None or value == '':
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed


def _validated_volume(item: dict[str, Any]) -> float | None:
    """Return the already-validated volume supplied by the common data contract."""
    validated = _as_float(item.get('validated_volume_m3'))
    if validated is not None:
        return max(validated, 0.0)
    if bool(item.get('period_m3_reliable')):
        reliable = _as_float(item.get('period_m3'))
        if reliable is not None:
            return max(reliable, 0.0)
    return None


def _volume_validation(item: dict[str, Any], validated: float | None = None) -> tuple[str, str]:
    quality_label = str(item.get('quality_label') or item.get('validation') or '').strip()
    quality_status = str(item.get('quality_status') or item.get('validation_status') or '').strip()
    if quality_label and quality_status:
        return quality_label, quality_status
    value = _validated_volume(item) if validated is None else validated
    if value is None:
        return 'Sin volumen validado', 'unavailable'
    return 'Validado', 'validated'


def _period_activity(item: dict[str, Any], validated: float | None = None) -> str:
    samples = int(item.get('samples_received') or item.get('samples') or 0)
    if samples <= 0:
        return 'Sin registros'
    active_samples = int(item.get('active_samples') or 0)
    source_activity = str(item.get('activity') or item.get('period_activity') or '').lower()
    value = _validated_volume(item) if validated is None else validated
    if active_samples > 0 or float(value or 0.0) > 0.0 or 'con actividad' in source_activity:
        return 'Con actividad'
    return 'Sin actividad'


def _report_row(item: dict[str, Any]) -> dict[str, Any]:
    closing_m3 = item.get('period_close_m3')
    if closing_m3 is None:
        closing_m3 = item.get('current_totalizer_m3')
    validated = _validated_volume(item)
    discarded = _as_float(item.get('discarded_volume_m3')) or 0.0
    validation, validation_status = _volume_validation(item, validated)
    reliable = bool(item.get('volume_reliable')) if 'volume_reliable' in item else bool(item.get('period_m3_reliable'))
    return {
        'operational_key': item.get('operational_key'),
        'sensor_id': item.get('sensor_id'),
        'module': item.get('module'),
        'name': item.get('name') or item.get('nombre'),
        'flow': item.get('current_flow'),
        'flow_unit': item.get('flow_unit') or 'L/s',
        'opening_m3': item.get('period_open_m3'),
        'closing_m3': closing_m3,
        'volume_m3': item.get('period_m3'),
        'volume_reliable': reliable,
        'validated_volume_m3': validated,
        'discarded_volume_m3': discarded,
        'discarded_totalizer_events': item.get('discarded_totalizer_events') or 0,
        'has_discontinuities': bool(item.get('has_discontinuities')),
        'volume_display_label': item.get('volume_display_label') or 'Volumen del periodo',
        'activity': item.get('activity') or _period_activity(item, validated),
        'validation': validation,
        'validation_status': validation_status,
        'quality_label': item.get('quality_label') or validation,
        'quality_status': item.get('quality_status') or validation_status,
        'quality_reason_code': item.get('quality_reason_code'),
        'quality_reason': item.get('quality_reason'),
        'quality_details': item.get('quality_details') or {},
        'communication': item.get('communication') or 'Sin lectura',
        'last_update': item.get('last_update'),
        'data_status': item.get('quality_data_status') or item.get('data_status') or 'no_data',
        'samples': item.get('samples_received') or item.get('samples') or 0,
        'samples_expected': item.get('samples_expected'),
        'coverage_percent': item.get('coverage_percent') or item.get('coverage_pct'),
        'opening_source': item.get('opening_source'),
        'boundary_complete': item.get('boundary_complete'),
    }


def _report_volume_display(item: dict[str, Any]) -> str:
    validated = _as_float(item.get('validated_volume_m3'))
    if validated is not None:
        return f'{_fmt_number(validated)} m³'
    return 'Sin volumen validado'


def _module_validated_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    values = [value for row in rows if (value := _validated_volume(row)) is not None]
    discarded = sum((_as_float(row.get('discarded_volume_m3')) or 0.0) for row in rows)
    statuses = [str(row.get('quality_status') or row.get('validation_status') or '') for row in rows]
    has_quality_contract = any(status in {'validated', 'valid_zero', 'partial_coverage', 'review', 'no_data'} for status in statuses)
    if has_quality_contract:
        partial_validation_count = sum(1 for status in statuses if status in {'partial_coverage', 'review'})
        without_validated_volume_count = sum(1 for status in statuses if status == 'no_data')
        coverage_complete = bool(rows) and all(status in {'validated', 'valid_zero'} for status in statuses)
    else:
        partial_validation_count = sum(1 for row in rows if _volume_validation(row)[1] == 'partial')
        without_validated_volume_count = sum(1 for row in rows if _volume_validation(row)[1] == 'unavailable')
        coverage_complete = bool(rows) and len(values) == len(rows)
    quality_counts: dict[str, int] = {}
    for status in statuses:
        key = status or 'unavailable'
        quality_counts[key] = quality_counts.get(key, 0) + 1
    return {
        'validated_volume_m3': round(sum(values), 6) if values else None,
        'discarded_volume_m3': round(discarded, 6),
        'calculable_count': len(values),
        'monitored_count': len(rows),
        'coverage_complete': coverage_complete,
        'active_count': sum(1 for row in rows if str(row.get('activity') or '').lower().startswith('con actividad')),
        'inactive_count': sum(1 for row in rows if str(row.get('activity') or '').lower().startswith('sin actividad')),
        'partial_validation_count': partial_validation_count,
        'without_validated_volume_count': without_validated_volume_count,
        'quality_counts': quality_counts,
        # Compatibility alias for clients from the previous report release.
        'review_count': partial_validation_count,
    }


def _history_aggregation(start_day: date, end_day: date) -> str:
    days = (end_day - start_day).days + 1
    if days == 1:
        return 'quarter_hour'
    if days <= 7:
        return 'hourly'
    return 'daily'


def _report_history(module: str, start_day: date, end_day: date, aggregation: str) -> dict[str, Any]:
    try:
        return get_water_history_module(
            module=module,
            start_date=start_day.isoformat(),
            end_date=end_day.isoformat(),
            aggregation=aggregation,
        )
    except (WaterHistoryError, ValueError):
        return {
            'module': module,
            'start_date': start_day.isoformat(),
            'end_date': end_day.isoformat(),
            'aggregation': aggregation,
            'series': [],
            'source_status': 'unavailable',
            'has_future_intervals': False,
        }


def _daily_review_as_period(day: date, *, include_shifts: bool) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Adapt the canonical daily-review payload to the report service without recalculating volumes."""
    try:
        review = get_daily_water_review(
            day.isoformat(),
            include_shifts=include_shifts,
            include_comparatives=False,
            force_refresh=False,
        )
    except DailyReviewError as exc:
        raise ReportDataUnavailableError(str(exc)) from exc

    modules = review.get('modules') or {}
    wells_module = modules.get('wells') or {}
    lines_module = modules.get('lines') or {}
    flows_module = modules.get('flows') or {}
    shifts_payload = review.get('shifts') or {}
    shifts = list(shifts_payload.get('shifts') or []) if isinstance(shifts_payload, dict) else []
    period = {
        'wells': list(wells_module.get('items') or review.get('wells') or []),
        'lines': list(lines_module.get('items') or review.get('production_lines') or []),
        'flows': list(flows_module.get('items') or review.get('flows') or []),
        'summary': {
            'wells': dict(wells_module.get('summary') or (review.get('operational_summary') or {}).get('wells') or {}),
            'lines': dict(lines_module.get('summary') or (review.get('operational_summary') or {}).get('lines') or {}),
            'flows': dict(flows_module.get('summary') or (review.get('operational_summary') or {}).get('flows') or {}),
        },
        'source_status': review.get('source_status'),
        'validated_segment_start': review.get('validated_segment_start'),
        'crosses_scada_cutover': bool(review.get('crosses_scada_cutover')),
        'legacy_notice': review.get('legacy_notice'),
        'has_future_intervals': bool(review.get('has_future_intervals')),
        'report_source': 'daily_review',
    }
    return period, shifts


def get_daily_water_report(
    report_date: Any = None,
    start_date: Any = None,
    end_date: Any = None,
    *,
    include_history: bool = True,
    include_shifts: bool = True,
) -> dict[str, Any]:
    start_day = _parse_date(start_date or report_date)
    end_day = _parse_date(end_date or report_date or start_day)
    if start_day > end_day:
        start_day, end_day = end_day, start_day
    if start_day == end_day:
        period, shifts = _daily_review_as_period(start_day, include_shifts=include_shifts)
    else:
        # Multi-day reports keep the existing reconciled period service. The canonical
        # daily-review contract is the source of truth for one-day reports.
        try:
            period = get_period_data(start_day.isoformat(), end_day.isoformat())
        except WaterPeriodError as exc:
            raise ReportDataUnavailableError(str(exc)) from exc
        period = dict(period)
        period['report_source'] = 'period_service'
        shifts = []

    wells = [_report_row(item) for item in period['wells']]
    lines = [_report_row(item) for item in period['lines']]
    flows = [_report_row(item) for item in period['flows']]
    summaries = period['summary']

    well_summary = _module_validated_summary(wells)
    line_summary = _module_validated_summary(lines)
    lavadoras, jarabes = _split_flow_rows(flows)
    lavadora_summary = _module_validated_summary(lavadoras)
    jarabes_summary = _module_validated_summary(jarabes)
    flow_summary = _module_validated_summary(flows)
    module_values = [
        well_summary['validated_volume_m3'],
        line_summary['validated_volume_m3'],
        lavadora_summary['validated_volume_m3'],
        jarabes_summary['validated_volume_m3'],
    ]
    calculable_values = [float(value) for value in module_values if value is not None]
    total_validated = round(sum(calculable_values), 6) if calculable_values else None
    partial_validation_count = sum(item['partial_validation_count'] for item in (well_summary, line_summary, lavadora_summary, jarabes_summary))
    without_validated_volume_count = sum(item['without_validated_volume_count'] for item in (well_summary, line_summary, lavadora_summary, jarabes_summary))
    monitored_items_count = sum(item['monitored_count'] for item in (well_summary, line_summary, lavadora_summary, jarabes_summary))
    coverage_complete = bool(monitored_items_count) and all(item['coverage_complete'] for item in (well_summary, line_summary, lavadora_summary, jarabes_summary))
    volume_basis_label = 'Total validado' if coverage_complete else 'Subtotal validado'
    latest = max((str(item.get('last_update') or '') for item in [*wells, *lines, *flows]), default='')
    period_label = start_day.strftime('%d/%m/%Y') if start_day == end_day else f'{start_day:%d/%m/%Y} al {end_day:%d/%m/%Y}'
    aggregation = _history_aggregation(start_day, end_day)
    history = {'aggregation': aggregation, 'wells': {}, 'lines': {}, 'flows': {}, 'washers': {}, 'jarabes': {}}
    if include_history:
        flow_history = _report_history('flow', start_day, end_day, aggregation)
        history.update({
            'wells': _report_history('well', start_day, end_day, aggregation),
            'lines': _report_history('line', start_day, end_day, aggregation),
            'flows': flow_history,
            'washers': _filter_history(flow_history, LAVADORA_KEYS),
            'jarabes': _filter_history(flow_history, JARABES_KEYS),
        })

    return {
        'title': 'Reporte Diario de Control Hídrico Durango',
        'plant': 'Planta Durango',
        'date': end_day.isoformat(),
        'start_date': start_day.isoformat(),
        'end_date': end_day.isoformat(),
        'period_label': period_label,
        'generated_at': datetime.now(LOCAL_ZONE).isoformat(timespec='seconds'),
        'source_status': period.get('source_status'),
        'report_source': period.get('report_source') or 'period_service',
        'summary': {
            # Backward-compatible fields consumed by the current dashboard.
            'well_volume_m3': well_summary['validated_volume_m3'],
            'line_volume_m3': line_summary['validated_volume_m3'],
            'flow_volume_m3': flow_summary['validated_volume_m3'],
            'washer_volume_m3': lavadora_summary['validated_volume_m3'],
            'jarabes_volume_m3': jarabes_summary['validated_volume_m3'],
            'total_operational_m3': total_validated,
            # Explicit fields for every report output.
            'well_validated_volume_m3': well_summary['validated_volume_m3'],
            'line_validated_volume_m3': line_summary['validated_volume_m3'],
            'flow_validated_volume_m3': flow_summary['validated_volume_m3'],
            'washer_validated_volume_m3': lavadora_summary['validated_volume_m3'],
            'jarabes_validated_volume_m3': jarabes_summary['validated_volume_m3'],
            'total_validated_operational_m3': total_validated,
            'discarded_volume_m3': round(
                well_summary['discarded_volume_m3']
                + line_summary['discarded_volume_m3']
                + lavadora_summary['discarded_volume_m3']
                + jarabes_summary['discarded_volume_m3'],
                6,
            ),
            'wells_active': summaries['wells']['active_count'],
            'lines_active': summaries['lines']['active_count'],
            'flows_active': summaries['flows']['active_count'],
            'washers_active': lavadora_summary['active_count'],
            'jarabes_active': jarabes_summary['active_count'],
            'partial_validation_count': partial_validation_count,
            'without_validated_volume_count': without_validated_volume_count,
            'validated_items_count': sum(1 for item in [*wells, *lines, *lavadoras, *jarabes] if item.get('validated_volume_m3') is not None),
            'monitored_items_count': monitored_items_count,
            'coverage_complete': coverage_complete,
            'coverage_label': 'Cobertura completa' if coverage_complete else 'Cobertura parcial',
            'volume_basis_label': volume_basis_label,
            'review_count': partial_validation_count,
            'no_data_count': without_validated_volume_count,
            'communication': 'Revisar comunicación' if any(item['communication'] != 'Actualizado' for item in [*wells, *lines, *flows]) else 'Actualizado',
            'last_update': latest or None,
            'note': (
                SUMMARY_NOTE
                if coverage_complete
                else SUMMARY_NOTE + ' El periodo contiene elementos con cobertura incompleta; el total se presenta como subtotal validado.'
            ),
        },
        'wells': {'rows': wells, **summaries['wells']},
        'production_lines': {'rows': lines, **summaries['lines']},
        'operational_flows': {'rows': flows, **summaries['flows']},
        'washers': {'rows': lavadoras, **lavadora_summary},
        'jarabes': {'rows': jarabes, **jarabes_summary},
        'validated_segment_start': period.get('validated_segment_start'),
        'crosses_scada_cutover': period.get('crosses_scada_cutover', False),
        'legacy_notice': period.get('legacy_notice'),
        'shifts': shifts,
        'shift_breakdown_available': bool(shifts),
        'history': history,
        'includes_history': include_history,
        'includes_shifts': include_shifts,
        'notes': [],
    }


def _logo_path() -> Path | None:
    candidates = [
        Path(__file__).resolve().parents[3] / 'frontend' / 'src' / 'assets' / 'arca-continental-logo.png',
        Path(__file__).resolve().parents[3] / 'frontend' / 'src' / 'assets' / 'arca-logo.png',
    ]
    return next((path for path in candidates if path.exists()), None)


def _pdf_table(rows: list[list[Any]], widths: list[float], *, repeat_rows: int = 1) -> Table:
    table = Table(rows, colWidths=widths, repeatRows=repeat_rows, hAlign='CENTER')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F1F8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#334155')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.1),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#C5D6E3')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAFC')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3.5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3.5),
        ('TOPPADDING', (0, 0), (-1, -1), 4.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4.5),
    ]))
    return table


PDF_COLORS = ['#1597D4', '#7047EB', '#F59E0B', '#10B981', '#E84A5F']


def _parse_history_timestamp(value: Any) -> datetime | None:
    try:
        return datetime.fromisoformat(str(value).replace('Z', '')).replace(tzinfo=None)
    except (TypeError, ValueError):
        return None


def _history_label(stamp: datetime, aggregation: str, single_day: bool) -> str:
    if aggregation == 'daily':
        return stamp.strftime('%d/%m')
    if single_day:
        return stamp.strftime('%H:%M')
    return stamp.strftime('%d/%m %Hh')


def _flow_history_drawing(history: dict[str, Any], *, width: float, height: float, single_day: bool) -> Drawing:
    drawing = Drawing(width, height)
    left, right, bottom, top = 42.0, 10.0, 42.0, 12.0
    plot_width = width - left - right
    plot_height = height - bottom - top
    drawing.add(Rect(left, bottom, plot_width, plot_height, fillColor=colors.HexColor('#FBFDFF'), strokeColor=colors.HexColor('#D7E4ED'), strokeWidth=0.6))
    series = list(history.get('series') or [])
    all_points = [point for item in series for point in item.get('points') or []]
    stamps = sorted({stamp for point in all_points if (stamp := _parse_history_timestamp(point.get('bucket_start') or point.get('timestamp'))) is not None})
    valid_values = [
        float(point['flow_avg_lps'])
        for point in all_points
        if int(point.get('samples') or 0) > 0 and point.get('flow_avg_lps') is not None
    ]
    if not stamps or not valid_values:
        drawing.add(String(width / 2, height / 2, 'Sin registros históricos para graficar', textAnchor='middle', fontName='Helvetica', fontSize=9, fillColor=colors.HexColor('#64748B')))
        return drawing

    start, end = stamps[0], stamps[-1]
    span = max((end - start).total_seconds(), 1.0)
    y_max = max(max(valid_values) * 1.12, 1.0)

    for index in range(5):
        ratio = index / 4
        y = bottom + plot_height * ratio
        drawing.add(Line(left, y, left + plot_width, y, strokeColor=colors.HexColor('#E4EDF3'), strokeWidth=0.45))
        drawing.add(String(left - 5, y - 2.5, f'{y_max * ratio:,.1f}', textAnchor='end', fontName='Helvetica', fontSize=6.5, fillColor=colors.HexColor('#64748B')))
    drawing.add(String(
        9,
        bottom + plot_height / 2,
        'L/s',
        angle=90,
        textAnchor='middle',
        fontName='Helvetica-Bold',
        fontSize=7,
        fillColor=colors.HexColor('#475569'),
    ))

    label_indexes = sorted({round(index * (len(stamps) - 1) / 4) for index in range(5)})
    aggregation = str(history.get('aggregation') or 'hourly')
    for index in label_indexes:
        stamp = stamps[index]
        x = left + ((stamp - start).total_seconds() / span) * plot_width
        drawing.add(Line(x, bottom, x, bottom - 3, strokeColor=colors.HexColor('#94A3B8'), strokeWidth=0.45))
        drawing.add(String(x, bottom - 12, _history_label(stamp, aggregation, single_day), textAnchor='middle', fontName='Helvetica', fontSize=6.2, fillColor=colors.HexColor('#64748B')))

    for series_index, item in enumerate(series):
        color = colors.HexColor(PDF_COLORS[series_index % len(PDF_COLORS)])
        values_by_stamp = {
            stamp: point
            for point in item.get('points') or []
            if (stamp := _parse_history_timestamp(point.get('bucket_start') or point.get('timestamp'))) is not None
        }
        path = DrawingPath()
        started = False
        has_segment = False
        for stamp in stamps:
            point = values_by_stamp.get(stamp)
            value = None if not point or int(point.get('samples') or 0) <= 0 else _as_float(point.get('flow_avg_lps'))
            if value is None:
                started = False
                continue
            x = left + ((stamp - start).total_seconds() / span) * plot_width
            y = bottom + (max(value, 0.0) / y_max) * plot_height
            if not started:
                path.moveTo(x, y)
                started = True
            else:
                path.lineTo(x, y)
            has_segment = True
        if has_segment:
            path.strokeColor = color
            path.strokeWidth = 1.55
            path.fillColor = None
            drawing.add(path)

        legend_column = series_index % 2
        legend_row = series_index // 2
        legend_x = left + legend_column * (plot_width / 2)
        legend_y = 8 + (1 - legend_row) * 10
        drawing.add(Line(legend_x, legend_y + 2, legend_x + 12, legend_y + 2, strokeColor=color, strokeWidth=2))
        drawing.add(String(legend_x + 16, legend_y, str(item.get('name') or 'Elemento'), fontName='Helvetica', fontSize=6.8, fillColor=colors.HexColor('#334155')))
    return drawing


def _validated_volume_drawing(rows: list[dict[str, Any]], *, width: float) -> Drawing:
    height = max(42 * mm, 16 * mm + len(rows) * 8 * mm)
    drawing = Drawing(width, height)
    label_width, value_width = 104.0, 72.0
    right = 8.0
    plot_width = width - label_width - value_width - right
    values = [_as_float(row.get('validated_volume_m3')) for row in rows]
    max_value = max((value for value in values if value is not None), default=0.0)
    scale_max = max(max_value, 1.0)
    row_height = (height - 18.0) / max(len(rows), 1)
    for index, row in enumerate(rows):
        y = height - 18.0 - index * row_height
        value = values[index]
        drawing.add(String(0, y + 2.0, str(row.get('name') or 'Elemento'), fontName='Helvetica', fontSize=7.2, fillColor=colors.HexColor('#334155')))
        drawing.add(Rect(label_width, y, plot_width, 8.0, fillColor=colors.HexColor('#EDF3F7'), strokeColor=None))
        if value is not None:
            color = '#1597D4'
            if value > 0:
                drawing.add(Rect(label_width, y, plot_width * value / scale_max, 8.0, fillColor=colors.HexColor(color), strokeColor=None))
            label = f'{value:,.2f} m³'
        else:
            label = 'Sin registros' if int(row.get('samples') or 0) <= 0 else 'Sin volumen validado'
        drawing.add(String(label_width + plot_width + 6, y + 1.5, label, fontName='Helvetica', fontSize=6.8, fillColor=colors.HexColor('#475569')))
    return drawing


def _report_footer(canvas, doc) -> None:
    canvas.saveState()
    width, _ = A4
    canvas.setStrokeColor(colors.HexColor('#D7E1E8'))
    canvas.setLineWidth(0.5)
    canvas.line(doc.leftMargin, 10 * mm, width - doc.rightMargin, 10 * mm)
    canvas.setFillColor(colors.HexColor('#64748B'))
    canvas.setFont('Helvetica', 7)
    canvas.drawString(doc.leftMargin, 6.5 * mm, 'Dashboard ARCA · Control hídrico · Planta Durango')
    canvas.drawRightString(width - doc.rightMargin, 6.5 * mm, f'Página {doc.page}')
    canvas.restoreState()


def build_daily_water_report_pdf(report: dict[str, Any]) -> tuple[bytes, str]:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=11 * mm,
        bottomMargin=16 * mm,
        title=str(report.get('title') or 'Reporte Diario de Control Hídrico Durango'),
        author='Dashboard ARCA',
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleDurango', parent=styles['Title'], alignment=TA_CENTER, fontSize=16, leading=19, textColor=colors.HexColor('#1F2937'), spaceAfter=2)
    eyebrow = ParagraphStyle('EyebrowDurango', parent=styles['BodyText'], alignment=TA_CENTER, fontSize=7.2, leading=9, textColor=colors.HexColor('#C8102E'), fontName='Helvetica-Bold', spaceAfter=2)
    heading = ParagraphStyle('HeadingDurango', parent=styles['Heading2'], fontSize=13, leading=16, textColor=colors.HexColor('#1F2937'), spaceBefore=2, spaceAfter=5)
    chart_heading = ParagraphStyle('ChartHeadingDurango', parent=styles['Heading3'], fontSize=9.5, leading=12, textColor=colors.HexColor('#334155'), spaceBefore=5, spaceAfter=4)
    small = ParagraphStyle('SmallDurango', parent=styles['BodyText'], fontSize=7.0, leading=8.8, textColor=colors.HexColor('#334155'))
    note_style = ParagraphStyle('NoteDurango', parent=small, fontSize=7.6, leading=10, textColor=colors.HexColor('#475569'))
    right = ParagraphStyle('RightDurango', parent=small, alignment=TA_RIGHT)
    center = ParagraphStyle('CenterDurango', parent=small, alignment=TA_CENTER)
    left = ParagraphStyle('LeftDurango', parent=small, alignment=TA_LEFT)
    story: list[Any] = []
    logo = _logo_path()
    if logo:
        logo_image = Image(str(logo), width=34 * mm, height=12 * mm, kind='proportional')
        logo_image.hAlign = 'CENTER'
        story.append(logo_image)
        story.append(Spacer(1, 1.5 * mm))
    story.append(Paragraph('DASHBOARD ARCA · PLANTA DURANGO', eyebrow))
    story.append(Paragraph('Reporte Diario de Control Hídrico', title_style))
    story.append(Paragraph(f"Periodo: {escape(str(report.get('period_label') or ''))} &nbsp;&nbsp;·&nbsp;&nbsp; Generado: {escape(_fmt_date(report.get('generated_at')))}", center))
    story.append(Spacer(1, 2.5 * mm))
    story.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#C8102E'), spaceAfter=4 * mm))

    summary = report['summary']
    story.append(Paragraph('Resumen ejecutivo', heading))
    kpis = [
        ('Volumen validado de pozos', _fmt_volume(summary.get('well_validated_volume_m3'))),
        ('Volumen validado de líneas', _fmt_volume(summary.get('line_validated_volume_m3'))),
        ('Volumen validado de lavadoras', _fmt_volume(summary.get('washer_validated_volume_m3'))),
        ('Volumen validado de Jarabes', _fmt_volume(summary.get('jarabes_validated_volume_m3'))),
        (f"{summary.get('volume_basis_label') or 'Total validado'} operativo", _fmt_volume(summary.get('total_validated_operational_m3'))),
        ('Pozos con actividad', f"{int(summary.get('wells_active') or 0)}/{len(report.get('wells', {}).get('rows', []))}"),
        ('Líneas con actividad', f"{int(summary.get('lines_active') or 0)}/{len(report.get('production_lines', {}).get('rows', []))}"),
        ('Lavadoras/Jarabes con actividad', f"{int(summary.get('washers_active') or 0) + int(summary.get('jarabes_active') or 0)}/{len(report.get('washers', {}).get('rows', [])) + len(report.get('jarabes', {}).get('rows', []))}"),
    ]
    cards = []
    for label, value in kpis:
        card = Table(
            [[Paragraph(escape(label.upper()), ParagraphStyle('KpiLabel', parent=center, fontSize=6.2, leading=7.5, textColor=colors.HexColor('#64748B'), fontName='Helvetica-Bold'))],
             [Paragraph(escape(value), ParagraphStyle('KpiValue', parent=center, fontSize=11.2, leading=13, textColor=colors.HexColor('#1F2937'), fontName='Helvetica-Bold'))]],
            colWidths=[44.5 * mm],
            rowHeights=[8 * mm, 10 * mm],
        )
        card.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
            ('BOX', (0, 0), (-1, -1), 0.55, colors.HexColor('#CBD9E4')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        cards.append(card)
    card_grid = Table([cards[:4], cards[4:]], colWidths=[46.5 * mm] * 4, rowHeights=[20 * mm, 20 * mm], hAlign='CENTER')
    card_grid.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('LEFTPADDING', (0, 0), (-1, -1), 1.5), ('RIGHTPADDING', (0, 0), (-1, -1), 1.5), ('TOPPADDING', (0, 0), (-1, -1), 1.5), ('BOTTOMPADDING', (0, 0), (-1, -1), 1.5)]))
    story.append(card_grid)
    story.append(Spacer(1, 3 * mm))
    quality_line = (
        f"<b>Calidad:</b> {escape(str(summary.get('coverage_label') or 'Sin dato'))}. "
        f"Elementos validados: {int(summary.get('validated_items_count') or 0)}/{int(summary.get('monitored_items_count') or 0)}. "
        f"En revisión: {int(summary.get('review_count') or 0)}. Sin datos: {int(summary.get('no_data_count') or 0)}."
    )
    note = Table([[Paragraph(
        escape(str(summary.get('note') or SUMMARY_NOTE)) + '<br/>' + quality_line + '<br/><b>Cero</b>: lectura válida sin flujo. <b>Hueco</b>: intervalo sin registros suficientes. Los gráficos no generan intervalos futuros.',
        note_style,
    )]], colWidths=[186 * mm])
    note.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F1F6FA')), ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#C7D8E4')), ('LEFTPADDING', (0, 0), (-1, -1), 8), ('RIGHTPADDING', (0, 0), (-1, -1), 8), ('TOPPADDING', (0, 0), (-1, -1), 7), ('BOTTOMPADDING', (0, 0), (-1, -1), 7)]))
    story.append(note)

    aggregation_label = {'quarter_hour': '15 minutos', 'hourly': '1 hora', 'daily': '1 día'}.get(str(report.get('history', {}).get('aggregation')), 'periodo')

    def section_block(title: str, section: dict[str, Any], history: dict[str, Any], first_name: str) -> list[Any]:
        data: list[list[Any]] = [[first_name, 'Flujo actual', 'Apertura', 'Cierre', 'Volumen validado', 'Actividad', 'Validación', 'Comunicación', 'Última actualización']]
        for item in section.get('rows', []):
            data.append([
                Paragraph(escape(str(item['name'])), left),
                Paragraph('No disponible' if item['flow'] is None else f"{_fmt_number(item['flow'])} {item['flow_unit']}", right),
                Paragraph(_fmt_volume(item['opening_m3']), right),
                Paragraph(_fmt_volume(item['closing_m3']), right),
                Paragraph(escape(_report_volume_display(item)), right),
                Paragraph(escape(str(item['activity'])), center),
                Paragraph(escape(str(item.get('validation') or 'Sin volumen validado')), center),
                Paragraph(escape(str(item['communication'])), center),
                Paragraph(escape(_fmt_date(item['last_update'])), center),
            ])
        table = _pdf_table(data, [18 * mm, 16 * mm, 19 * mm, 19 * mm, 22 * mm, 19 * mm, 20 * mm, 20 * mm, 33 * mm])
        table.setStyle(TableStyle([
            ('ALIGN', (1, 1), (4, -1), 'RIGHT'),
            ('ALIGN', (5, 1), (-1, -1), 'CENTER'),
        ]))
        return [
            Paragraph(title, heading),
            Paragraph(f"Periodo {escape(str(report.get('period_label') or ''))} · Agrupación histórica: {aggregation_label}", small),
            Spacer(1, 2 * mm),
            table,
            Paragraph(f'Comportamiento de flujo · {title}', chart_heading),
            _flow_history_drawing(history, width=186 * mm, height=64 * mm, single_day=report.get('start_date') == report.get('end_date')),
            Paragraph('Volumen validado por elemento', chart_heading),
            _validated_volume_drawing(section.get('rows', []), width=186 * mm),
        ]

    report_history = report.get('history') or {}
    for title, section, history_key, first_name in [
        ('Pozos', report['wells'], 'wells', 'Pozo'),
        ('Líneas', report['production_lines'], 'lines', 'Línea'),
        ('Lavadoras', report['washers'], 'washers', 'Lavadora'),
        ('Jarabes', report['jarabes'], 'jarabes', 'Jarabes'),
    ]:
        story.append(PageBreak())
        story.append(KeepTogether(section_block(title, section, report_history.get(history_key) or {}, first_name)))

    if report.get('shifts'):
        story.append(PageBreak())
        story.append(Paragraph('Cortes por turno', heading))
        story.append(Paragraph('Cortes administrativos calculados con las mismas lecturas normalizadas del dashboard.', small))
        story.append(Spacer(1, 2 * mm))
        shift_rows = [['Turno', 'Horario', 'Pozos', 'Líneas', 'Lavadoras', 'Jarabes', 'Total operativo', 'Estado']]
        for shift in report['shifts']:
            summary_shift = shift.get('summary') or {}
            shift_rows.append([
                shift.get('name'),
                shift.get('schedule'),
                _fmt_volume((summary_shift.get('wells') or {}).get('total_m3')),
                _fmt_volume((summary_shift.get('lines') or {}).get('total_m3')),
                _fmt_volume(_module_validated_summary([item for item in shift.get('flows') or [] if _is_lavadora(item)])['validated_volume_m3']),
                _fmt_volume(_module_validated_summary([item for item in shift.get('flows') or [] if _is_jarabes(item)])['validated_volume_m3']),
                _fmt_volume(summary_shift.get('total_operational_m3')), 
                shift.get('cut_status'),
            ])
        story.append(_pdf_table(shift_rows, [20 * mm, 24 * mm, 23 * mm, 23 * mm, 25 * mm, 23 * mm, 29 * mm, 19 * mm]))

    doc.build(story, onFirstPage=_report_footer, onLaterPages=_report_footer)
    filename = f"reporte-diario-control-hidrico-durango-{report.get('start_date')}.pdf"
    return buffer.getvalue(), filename


def _style_sheet(ws) -> None:
    header_fill = PatternFill('solid', fgColor='0B6E8E')
    alternate_fill = PatternFill('solid', fgColor='F3F8FA')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '0B6E8E'
    for row_index, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if row_index % 2 == 0:
                cell.fill = alternate_fill
        ws.row_dimensions[row_index].height = 22
    for column in range(1, ws.max_column + 1):
        width = max(len(str(ws.cell(row=row, column=column).value or '')) for row in range(1, ws.max_row + 1)) + 2
        ws.column_dimensions[get_column_letter(column)].width = min(max(width, 12), 32)


def _excel_datetime(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace('Z', '')).replace(tzinfo=None)
    except ValueError:
        return None


def build_daily_water_report_excel(report: dict[str, Any]) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen'
    ws.append(['Concepto', 'Valor'])
    summary = report['summary']
    for label, value in [
        ('Planta', report['plant']),
        ('Periodo', report['period_label']),
        ('Fecha de generación', datetime.fromisoformat(report['generated_at']).replace(tzinfo=None)),
        ('Volumen validado de pozos (m³)', summary['well_validated_volume_m3']),
        ('Volumen validado de líneas (m³)', summary['line_validated_volume_m3']),
        ('Volumen validado de lavadoras (m³)', summary['washer_validated_volume_m3']),
        ('Volumen validado de Jarabes (m³)', summary['jarabes_validated_volume_m3']),
        (f"{summary.get('volume_basis_label') or 'Total validado'} operativo (m³)", summary['total_validated_operational_m3']),
        ('Elementos validados', summary.get('validated_items_count', 0)),
        ('Elementos monitoreados', summary.get('monitored_items_count', 0)),
        ('Cobertura del reporte', summary.get('coverage_label')),
        ('Elementos en revisión', summary.get('review_count', 0)),
        ('Elementos sin datos', summary.get('no_data_count', 0)),
        ('Fuente de datos del reporte', 'Revisión diaria conciliada' if report.get('report_source') == 'daily_review' else 'Periodo conciliado'),
        ('Estado de comunicación', summary['communication']),
        ('Criterio de cálculo', summary.get('note') or SUMMARY_NOTE),
    ]:
        ws.append([label, value])
    ws['B4'].number_format = 'dd/mm/yyyy hh:mm'
    for row in range(5, 9):
        if isinstance(ws.cell(row, 2).value, (int, float)):
            ws.cell(row, 2).number_format = '#,##0.00'
    _style_sheet(ws)
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 72
    ws.row_dimensions[17].height = 48

    def add_items_sheet(name: str, rows: list[dict[str, Any]]) -> None:
        sheet = wb.create_sheet(name)
        sheet.append([
            'Elemento',
            'Flujo actual (L/s)',
            'Apertura (m³)',
            'Cierre (m³)',
            'Volumen validado (m³)',
            'Estado de validación',
            'Actividad',
            'Comunicación',
            'Última actualización',
        ])
        for item in rows:
            sheet.append([
                item['name'],
                item['flow'],
                item['opening_m3'],
                item['closing_m3'],
                item['validated_volume_m3'],
                item.get('validation') or _volume_validation(item)[0],
                item['activity'],
                item['communication'],
                _excel_datetime(item.get('last_update')),
            ])
        for row in range(2, sheet.max_row + 1):
            for column in (2, 3, 4, 5):
                if isinstance(sheet.cell(row, column).value, (int, float)):
                    sheet.cell(row, column).number_format = '#,##0.00'
            if isinstance(sheet.cell(row, 9).value, datetime):
                sheet.cell(row, 9).number_format = 'dd/mm/yyyy hh:mm'
        _style_sheet(sheet)

    add_items_sheet('Pozos', report['wells']['rows'])
    add_items_sheet('Líneas', report['production_lines']['rows'])
    add_items_sheet('Lavadoras', report['washers']['rows'])
    add_items_sheet('Jarabes', report['jarabes']['rows'])

    shifts = wb.create_sheet('Turnos')
    shifts.append(['Turno', 'Horario', 'Pozos (m³)', 'Líneas (m³)', 'Lavadoras (m³)', 'Jarabes (m³)', 'Total operativo (m³)', 'Estado'])
    for shift in report.get('shifts') or []:
        shift_summary = shift.get('summary') or {}
        shifts.append([
            shift.get('name'),
            shift.get('schedule'),
            (shift_summary.get('wells') or {}).get('total_m3'),
            (shift_summary.get('lines') or {}).get('total_m3'),
            _module_validated_summary([item for item in shift.get('flows') or [] if _is_lavadora(item)])['validated_volume_m3'],
            _module_validated_summary([item for item in shift.get('flows') or [] if _is_jarabes(item)])['validated_volume_m3'],
            shift_summary.get('total_operational_m3'),
            shift.get('cut_status'),
        ])
    for row in range(2, shifts.max_row + 1):
        for column in range(3, 8):
            if isinstance(shifts.cell(row, column).value, (int, float)):
                shifts.cell(row, column).number_format = '#,##0.00'
    _style_sheet(shifts)

    def add_history_sheet(name: str, history: dict[str, Any]) -> None:
        sheet = wb.create_sheet(name)
        sheet.append([
            'Elemento',
            'Inicio del intervalo',
            'Fin del intervalo',
            'Agrupación',
            'Flujo promedio del intervalo (L/s)',
            'Flujo promedio activo (L/s)',
            'Flujo mínimo (L/s)',
            'Flujo máximo (L/s)',
            'Minutos activos',
            'Muestras recibidas',
            'Muestras esperadas',
            'Cobertura (%)',
            'Totalizador apertura (m³)',
            'Totalizador cierre (m³)',
            'Volumen validado (m³)',
            'Estado del intervalo',
            'Estado de datos',
        ])
        for series in history.get('series') or []:
            for point in series.get('points') or []:
                start = _parse_history_timestamp(point.get('bucket_start') or point.get('timestamp'))
                end = _parse_history_timestamp(point.get('bucket_end'))
                sheet.append([
                    series.get('name'),
                    start,
                    end,
                    point.get('aggregation') or history.get('aggregation'),
                    point.get('flow_avg_lps'),
                    point.get('flow_active_avg_lps'),
                    point.get('flow_min_lps'),
                    point.get('flow_max_lps'),
                    point.get('active_minutes'),
                    point.get('samples_received', point.get('samples')),
                    point.get('samples_expected'),
                    point.get('coverage_percent'),
                    point.get('totalizer_open_m3'),
                    point.get('totalizer_close_m3'),
                    point.get('validated_volume_m3'),
                    point.get('interval_state'),
                    point.get('data_status'),
                ])
        for row in range(2, sheet.max_row + 1):
            for column in (2, 3):
                if isinstance(sheet.cell(row, column).value, datetime):
                    sheet.cell(row, column).number_format = 'dd/mm/yyyy hh:mm'
            for column in range(5, 16):
                if isinstance(sheet.cell(row, column).value, (int, float)):
                    sheet.cell(row, column).number_format = '#,##0.00'
        _style_sheet(sheet)

    # Historical sheets are complementary annexes behind the PDF charts.
    report_history = report.get('history') or {}
    add_history_sheet('Histórico Pozos', report_history.get('wells') or {})
    add_history_sheet('Histórico Líneas', report_history.get('lines') or {})
    add_history_sheet('Histórico Lavadoras', report_history.get('washers') or {})
    add_history_sheet('Histórico Jarabes', report_history.get('jarabes') or {})

    if report.get('shifts'):
        detail = wb.create_sheet('Detalle turnos')
        detail.append(['Turno', 'Grupo', 'Elemento', 'Apertura (m³)', 'Cierre (m³)', 'Volumen validado (m³)', 'Validación', 'Flujo promedio (L/s)', 'Flujo mínimo (L/s)', 'Flujo máximo (L/s)', 'Muestras', 'Actividad', 'Estado'])
        for shift in report['shifts']:
            for group_key, group_name in [('wells', 'Pozos'), ('lines', 'Líneas')]:
                for item in shift.get(group_key) or []:
                    validation, _ = _volume_validation(item)
                    detail.append([
                        shift.get('name'),
                        group_name,
                        item.get('name'),
                        item.get('period_open_m3'),
                        item.get('period_close_m3'),
                        item.get('validated_volume_m3'),
                        validation,
                        item.get('flow_avg'),
                        item.get('flow_min'),
                        item.get('flow_max'),
                        item.get('samples'),
                        item.get('activity'),
                        shift.get('cut_status'),
                    ])
            for group_name, allowed in [('Lavadoras', LAVADORA_KEYS), ('Jarabes', JARABES_KEYS)]:
                for item in [flow for flow in shift.get('flows') or [] if _operational_key(flow) in allowed]:
                    validation, _ = _volume_validation(item)
                    detail.append([
                        shift.get('name'),
                        group_name,
                        item.get('name'),
                        item.get('period_open_m3'),
                        item.get('period_close_m3'),
                        item.get('validated_volume_m3'),
                        validation,
                        item.get('flow_avg'),
                        item.get('flow_min'),
                        item.get('flow_max'),
                        item.get('samples'),
                        item.get('activity'),
                        shift.get('cut_status'),
                    ])
        for row in range(2, detail.max_row + 1):
            for column in (4, 5, 6, 8, 9, 10):
                if isinstance(detail.cell(row, column).value, (int, float)):
                    detail.cell(row, column).number_format = '#,##0.00'
        _style_sheet(detail)

    output = BytesIO()
    wb.save(output)
    filename = f"reporte-diario-control-hidrico-durango-{report.get('start_date')}.xlsx"
    return output.getvalue(), filename
