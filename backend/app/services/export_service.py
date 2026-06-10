from __future__ import annotations

import base64
import io
import re
import tempfile
from pathlib import Path
from uuid import uuid4

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

from app.schemas.dashboard import DashboardPayload

EXPORT_DIR = Path('exports')
EXPORT_DIR.mkdir(exist_ok=True)

EXCEL_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

THEME = {
    'bg': '#060916',
    'panel': '#0f1728',
    'panel_alt': '#101828',
    'grid': '#24314c',
    'text': '#f8fbff',
    'muted': '#c9d5ea',
    'red_1': '#e9082c',
    'red_2': '#c40323',
    'red_3': '#a0021c',
    'brown': '#8a3d26',
    'soft': '#f6c2cb',
}


def _safe_name(name: str) -> str:
    normalized = re.sub(r'[^a-zA-Z0-9_-]+', '_', name.strip().lower())
    normalized = re.sub(r'_+', '_', normalized).strip('_')
    return normalized or 'reporte'


def _target(payload: DashboardPayload, suffix: str) -> Path:
    return EXPORT_DIR / f'{_safe_name(payload.title)}_{uuid4().hex[:8]}.{suffix}'


def payload_to_dataframe(payload: DashboardPayload) -> pd.DataFrame:
    frame = pd.DataFrame([row.model_dump() for row in payload.table_data])
    if not frame.empty and 'timestamp' in frame.columns:
        frame['timestamp'] = pd.to_datetime(frame['timestamp']).dt.strftime('%Y-%m-%d %H:%M:%S')
    return frame


def _style_worksheet(ws, title: str):
    header_fill = PatternFill(fill_type='solid', start_color='A0021C', end_color='A0021C')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='263252')

    ws.insert_rows(1, 2)
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(fill_type='solid', start_color='0F172A', end_color='0F172A')
    ws['A2'] = 'ARCA CONTINENTAL · Reporte ejecutivo'
    ws['A2'].font = Font(size=10, italic=True, color='D9E2F1')
    ws['A2'].fill = PatternFill(fill_type='solid', start_color='0F172A', end_color='0F172A')

    for row in ws.iter_rows(min_row=3, max_row=3):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    for column in ws.columns:
        values = [str(cell.value or '') for cell in column]
        max_len = max((len(v) for v in values), default=10)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_len + 2, 12), 42)

    ws.freeze_panes = 'A4'
    ws.sheet_view.showGridLines = False


def _save_figure(fig) -> bytes:
    buffer = io.BytesIO()
    fig.savefig(buffer, dpi=180, bbox_inches='tight', facecolor=fig.get_facecolor())
    plt.close(fig)
    buffer.seek(0)
    return buffer.getvalue()


def _make_line_chart(payload: DashboardPayload) -> bytes:
    data = payload.hourly_data
    fig, ax = plt.subplots(figsize=(9.2, 4.2), facecolor=THEME['bg'])
    ax.set_facecolor(THEME['panel'])
    labels = [item.hour for item in data]
    totals = [item.total for item in data]
    l1 = [item.l1 for item in data]
    l2 = [item.l2 for item in data]
    l3 = [item.l3 for item in data]
    ax.plot(labels, totals, color=THEME['red_1'], linewidth=2.8, label='Total')
    ax.fill_between(labels, totals, color=THEME['red_1'], alpha=0.16)
    ax.plot(labels, l1, color=THEME['red_2'], linewidth=2.0, label='L1')
    ax.plot(labels, l2, color=THEME['red_3'], linewidth=2.0, label='L2')
    ax.plot(labels, l3, color=THEME['brown'], linewidth=2.0, label='L3')
    ax.grid(color=THEME['grid'], linestyle='--', alpha=0.55)
    ax.tick_params(axis='x', colors=THEME['muted'], rotation=35)
    ax.tick_params(axis='y', colors=THEME['muted'])
    for spine in ax.spines.values():
        spine.set_color('#334155')
    ax.set_title('Demanda por hora', color=THEME['text'], fontsize=16, pad=10)
    ax.set_ylabel('kW', color=THEME['muted'])
    ax.legend(facecolor=THEME['panel'], edgecolor=THEME['grid'], labelcolor=THEME['text'])
    fig.tight_layout()
    return _save_figure(fig)


def _make_distribution_chart(payload: DashboardPayload) -> bytes:
    labels = [item.name for item in payload.systems_data[:8]]
    values = [item.value for item in payload.systems_data[:8]]
    colorset = [THEME['red_1'], THEME['red_2'], THEME['red_3'], THEME['brown'], '#f04a62', '#ff9078', '#f6c2cb', '#b15a3c']
    fig, ax = plt.subplots(figsize=(7.2, 4.2), facecolor=THEME['bg'])
    ax.set_facecolor(THEME['panel'])
    wedges, _ = ax.pie(values, colors=colorset[:len(values)], startangle=90, wedgeprops={'width': 0.42, 'edgecolor': THEME['bg']})
    ax.legend(wedges, labels, loc='center left', bbox_to_anchor=(1, 0.5), frameon=False, labelcolor=THEME['text'])
    ax.set_title('Distribución de consumo', color=THEME['text'], fontsize=16, pad=10)
    fig.tight_layout()
    return _save_figure(fig)


def _make_transformer_chart(payload: DashboardPayload) -> bytes:
    labels = [item.name for item in payload.transformer_data]
    values = [item.kwh for item in payload.transformer_data]
    fig, ax = plt.subplots(figsize=(8.4, 4.0), facecolor=THEME['bg'])
    ax.set_facecolor(THEME['panel'])
    bars = ax.bar(labels, values, color=[THEME['red_1'], THEME['red_2'], THEME['red_3'], THEME['brown'], '#f04a62'][:len(labels)], edgecolor='none')
    ax.grid(axis='y', color=THEME['grid'], linestyle='--', alpha=0.55)
    ax.tick_params(axis='x', colors=THEME['muted'], rotation=0)
    ax.tick_params(axis='y', colors=THEME['muted'])
    ax.set_title('Transformadores por energía', color=THEME['text'], fontsize=16, pad=10)
    ax.set_ylabel('kWh', color=THEME['muted'])
    for spine in ax.spines.values():
        spine.set_color('#334155')
    for b in bars:
        ax.text(b.get_x()+b.get_width()/2, b.get_height(), f'{b.get_height():.0f}', ha='center', va='bottom', color=THEME['text'], fontsize=9)
    fig.tight_layout()
    return _save_figure(fig)


def _save_temp_image(image_bytes: bytes) -> str:
    temp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
    temp.write(image_bytes)
    temp.flush()
    temp.close()
    return temp.name


def export_excel(payload: DashboardPayload) -> Path:
    path = _target(payload, 'xlsx')
    cards_df = pd.DataFrame([card.model_dump() for card in payload.cards])
    hourly_df = pd.DataFrame([item.model_dump() for item in payload.hourly_data])
    distribution_df = pd.DataFrame([item.model_dump() for item in payload.systems_data])
    transformer_df = pd.DataFrame([item.model_dump() for item in payload.transformer_data])
    data_df = payload_to_dataframe(payload)
    summary_df = pd.DataFrame([
        {'Título': payload.title, 'Subtítulo': payload.subtitle, 'Actualizado': payload.updated_at.strftime('%Y-%m-%d %H:%M:%S')},
    ])

    line_img = _save_temp_image(_make_line_chart(payload))
    dist_img = _save_temp_image(_make_distribution_chart(payload))
    tr_img = _save_temp_image(_make_transformer_chart(payload))

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Resumen', index=False, startrow=2)
        cards_df.to_excel(writer, sheet_name='KPIs', index=False, startrow=2)
        hourly_df.to_excel(writer, sheet_name='Historico', index=False, startrow=2)
        distribution_df.to_excel(writer, sheet_name='Distribucion', index=False, startrow=2)
        transformer_df.to_excel(writer, sheet_name='Transformadores', index=False, startrow=2)
        data_df.to_excel(writer, sheet_name='Mediciones', index=False, startrow=2)

        for sheet_name in ['Resumen', 'KPIs', 'Historico', 'Distribucion', 'Transformadores', 'Mediciones']:
            _style_worksheet(writer.book[sheet_name], f'{payload.title} · {sheet_name}')

        chart_sheet = writer.book.create_sheet('Graficas')
        chart_sheet['A1'] = f'{payload.title} · Gráficas'
        chart_sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        chart_sheet.sheet_view.showGridLines = False
        chart_sheet['A2'] = 'Las gráficas se incluyen para impresión y revisión ejecutiva.'
        chart_sheet['A2'].font = Font(size=10, italic=True, color='D9E2F1')
        chart_sheet.add_image(XLImage(line_img), 'A4')
        chart_sheet.add_image(XLImage(dist_img), 'A28')
        chart_sheet.add_image(XLImage(tr_img), 'J4')

    return path


def export_csv(payload: DashboardPayload) -> Path:
    path = _target(payload, 'csv')
    payload_to_dataframe(payload).to_csv(path, index=False)
    return path


def export_html(payload: DashboardPayload) -> Path:
    path = _target(payload, 'html')
    cards_html = ''.join([
        f"<div class='kpi-card'><div class='kpi-label'>{c.label}</div><div class='kpi-value'>{c.value} <span>{c.unit}</span></div><div class='kpi-trend'>{c.trend}</div></div>"
        for c in payload.cards
    ])
    rows = ''.join([
        f"<tr><td>{row.timestamp:%Y-%m-%d %H:%M}</td><td>{row.system_name}</td><td>{row.kw:.2f}</td><td>{row.kwh:.2f}</td><td>{row.kvarh:.2f}</td><td>{row.current:.2f}</td><td>{row.power_factor:.3f}</td><td>{str(row.status).upper()}</td></tr>"
        for row in payload.table_data
    ])
    line_b64 = base64.b64encode(_make_line_chart(payload)).decode('ascii')
    dist_b64 = base64.b64encode(_make_distribution_chart(payload)).decode('ascii')
    tr_b64 = base64.b64encode(_make_transformer_chart(payload)).decode('ascii')
    html = f"""
    <html>
    <head>
      <meta charset='utf-8'>
      <title>{payload.title}</title>
      <style>
        body {{ font-family: Inter, Arial, sans-serif; background:#060916; color:#f8fbff; padding:32px; }}
        .shell {{ max-width: 1180px; margin:0 auto; }}
        .hero {{ background: linear-gradient(135deg, rgba(233,8,44,.16), rgba(255,255,255,.03)); border:1px solid rgba(233,8,44,.18); border-radius:24px; padding:28px; margin-bottom:22px; }}
        .eyebrow {{ color:#ffccd5; font-size:12px; text-transform:uppercase; letter-spacing:.18em; }}
        h1 {{ margin:10px 0 4px; font-size:34px; }}
        p {{ color:#b9c6dc; }}
        .grid {{ display:grid; grid-template-columns:repeat(4,1fr); gap:16px; margin:20px 0 24px; }}
        .charts {{ display:grid; grid-template-columns:1.35fr 1fr; gap:16px; margin:0 0 24px; }}
        .chart-card {{ background:#0f1728; border:1px solid rgba(233,8,44,.14); border-radius:18px; padding:18px; }}
        .chart-card img {{ width:100%; border-radius:14px; display:block; }}
        .kpi-card {{ background:#0f1728; border:1px solid rgba(233,8,44,.14); border-radius:18px; padding:18px; }}
        .kpi-label {{ font-size:12px; text-transform:uppercase; letter-spacing:.14em; color:#c9d5ea; }}
        .kpi-value {{ font-size:30px; font-weight:800; margin-top:10px; }}
        .kpi-value span {{ font-size:15px; color:#dfe7f5; }}
        .kpi-trend {{ color:#b9c6dc; margin-top:8px; font-size:14px; }}
        table {{ width:100%; border-collapse:collapse; background:#0f1728; border-radius:18px; overflow:hidden; }}
        th {{ background:#a0021c; color:#fff; text-transform:uppercase; letter-spacing:.12em; font-size:11px; }}
        th, td {{ padding:12px 10px; border-bottom:1px solid #24314c; text-align:left; }}
        td:nth-last-child(-n+5), th:nth-last-child(-n+5) {{ text-align:right; }}
      </style>
    </head>
    <body>
      <div class='shell'>
        <section class='hero'>
          <div class='eyebrow'>ARCA CONTINENTAL</div>
          <h1>{payload.title}</h1>
          <p>{payload.subtitle} · Actualizado {payload.updated_at:%Y-%m-%d %H:%M:%S}</p>
        </section>
        <section class='grid'>{cards_html}</section>
        <section class='charts'>
          <div class='chart-card'><h3>Demanda por hora</h3><img src='data:image/png;base64,{line_b64}' /></div>
          <div class='chart-card'><h3>Distribución de consumo</h3><img src='data:image/png;base64,{dist_b64}' /></div>
        </section>
        <section class='charts'>
          <div class='chart-card'><h3>Transformadores por energía</h3><img src='data:image/png;base64,{tr_b64}' /></div>
          <div class='chart-card'><h3>Resumen ejecutivo</h3><p>Las gráficas se incluyen junto con KPIs y mediciones para impresión, envío y análisis rápido.</p></div>
        </section>
        <section>
          <h2>Mediciones recientes</h2>
          <table>
            <thead><tr><th>Fecha</th><th>Sistema</th><th>kW</th><th>kWh</th><th>kVARh</th><th>Corriente</th><th>FP</th><th>Estado</th></tr></thead>
            <tbody>{rows}</tbody>
          </table>
        </section>
      </div>
    </body>
    </html>
    """
    path.write_text(html, encoding='utf-8')
    return path


def export_pdf(payload: DashboardPayload) -> Path:
    path = _target(payload, 'pdf')
    pdf = canvas.Canvas(str(path), pagesize=A4)
    width, height = A4
    line_img = ImageReader(io.BytesIO(_make_line_chart(payload)))
    dist_img = ImageReader(io.BytesIO(_make_distribution_chart(payload)))
    tr_img = ImageReader(io.BytesIO(_make_transformer_chart(payload)))

    pdf.setFillColor(colors.HexColor('#0F172A'))
    pdf.roundRect(14 * mm, height - 50 * mm, width - 28 * mm, 34 * mm, 8 * mm, stroke=0, fill=1)
    pdf.setStrokeColor(colors.HexColor('#A0021C'))
    pdf.roundRect(14 * mm, height - 50 * mm, width - 28 * mm, 34 * mm, 8 * mm, stroke=1, fill=0)
    pdf.setFillColor(colors.white)
    pdf.setFont('Helvetica-Bold', 20)
    pdf.drawString(22 * mm, height - 28 * mm, payload.title)
    pdf.setFont('Helvetica', 10)
    pdf.drawString(22 * mm, height - 34 * mm, payload.subtitle)
    pdf.drawRightString(width - 22 * mm, height - 34 * mm, payload.updated_at.strftime('%Y-%m-%d %H:%M:%S'))

    card_y = height - 66 * mm
    card_w = (width - 40 * mm) / 2
    for index, card in enumerate(payload.cards[:4]):
        x = 20 * mm + (index % 2) * (card_w + 6 * mm)
        y = card_y - (index // 2) * 24 * mm
        pdf.setFillColor(colors.HexColor('#101828'))
        pdf.roundRect(x, y, card_w, 20 * mm, 5 * mm, stroke=0, fill=1)
        pdf.setStrokeColor(colors.HexColor('#C40323'))
        pdf.roundRect(x, y, card_w, 20 * mm, 5 * mm, stroke=1, fill=0)
        pdf.setFillColor(colors.HexColor('#C9D5EA'))
        pdf.setFont('Helvetica-Bold', 9)
        pdf.drawString(x + 4 * mm, y + 15 * mm, card.label.upper())
        pdf.setFillColor(colors.white)
        pdf.setFont('Helvetica-Bold', 18)
        pdf.drawString(x + 4 * mm, y + 8 * mm, f'{card.value}')
        pdf.setFont('Helvetica', 10)
        pdf.drawString(x + 35 * mm, y + 8 * mm, card.unit)
        pdf.setFillColor(colors.HexColor('#B9C6DC'))
        pdf.drawString(x + 4 * mm, y + 3 * mm, card.trend)

    pdf.drawImage(line_img, 20 * mm, height - 176 * mm, width=108 * mm, height=52 * mm, mask='auto')
    pdf.drawImage(dist_img, 132 * mm, height - 176 * mm, width=58 * mm, height=52 * mm, mask='auto')
    pdf.drawImage(tr_img, 20 * mm, height - 232 * mm, width=170 * mm, height=42 * mm, mask='auto')

    table_y = height - 246 * mm
    pdf.setFillColor(colors.white)
    pdf.setFont('Helvetica-Bold', 12)
    pdf.drawString(20 * mm, table_y, 'Mediciones recientes')
    table_y -= 6 * mm
    headers = ['Fecha', 'Sistema', 'kW', 'kWh', 'kVARh', 'FP', 'Estado']
    widths = [34 * mm, 46 * mm, 16 * mm, 18 * mm, 18 * mm, 14 * mm, 22 * mm]
    x = 20 * mm
    pdf.setFillColor(colors.HexColor('#A0021C'))
    pdf.rect(x, table_y - 6 * mm, sum(widths), 8 * mm, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont('Helvetica-Bold', 8)
    for header, w in zip(headers, widths):
        pdf.drawString(x + 1.5 * mm, table_y - 3.3 * mm, header)
        x += w

    y = table_y - 11 * mm
    pdf.setFont('Helvetica', 7.6)
    for row in payload.table_data[:14]:
        values = [
            row.timestamp.strftime('%Y-%m-%d %H:%M'),
            row.system_name[:28],
            f'{row.kw:.2f}',
            f'{row.kwh:.2f}',
            f'{row.kvarh:.2f}',
            f'{row.power_factor:.3f}',
            str(row.status).upper(),
        ]
        x = 20 * mm
        for value, w in zip(values, widths):
            pdf.setFillColor(colors.HexColor('#E8EEF8'))
            pdf.drawString(x + 1.5 * mm, y, value)
            pdf.setStrokeColor(colors.HexColor('#24314C'))
            pdf.line(x, y - 1.8 * mm, x + w, y - 1.8 * mm)
            x += w
        y -= 6 * mm

    pdf.save()
    return path


def export_png(payload: DashboardPayload) -> Path:
    path = _target(payload, 'png')
    fig = plt.figure(figsize=(13.5, 8.2), facecolor=THEME['bg'])
    gs = fig.add_gridspec(2, 2, height_ratios=[1, 1.15], width_ratios=[1.4, 1])
    ax1 = fig.add_subplot(gs[0, 0])
    ax2 = fig.add_subplot(gs[0, 1])
    ax3 = fig.add_subplot(gs[1, :])

    for ax in [ax1, ax2, ax3]:
        ax.set_facecolor(THEME['panel'])
        for spine in ax.spines.values():
            spine.set_color('#334155')
        ax.tick_params(colors=THEME['muted'])
        ax.grid(color=THEME['grid'], linestyle='--', alpha=0.45)

    labels = [item.hour for item in payload.hourly_data]
    totals = [item.total for item in payload.hourly_data]
    ax1.plot(labels, totals, color=THEME['red_1'], linewidth=2.8)
    ax1.fill_between(labels, totals, color=THEME['red_1'], alpha=0.15)
    ax1.set_title('Demanda por hora', color=THEME['text'], fontsize=16)
    ax1.tick_params(axis='x', rotation=35)

    dlabels = [item.name for item in payload.systems_data[:6]]
    dvalues = [item.value for item in payload.systems_data[:6]]
    ax2.barh(dlabels, dvalues, color=[THEME['red_1'], THEME['red_2'], THEME['red_3'], THEME['brown'], '#f04a62', '#ff9078'][:len(dlabels)])
    ax2.set_title('Distribución de consumo', color=THEME['text'], fontsize=16)

    tlabels = [item.name for item in payload.transformer_data]
    tvalues = [item.kwh for item in payload.transformer_data]
    ax3.bar(tlabels, tvalues, color=[THEME['red_1'], THEME['red_2'], THEME['red_3'], THEME['brown'], '#f04a62'][:len(tlabels)])
    ax3.set_title('Transformadores por energía (kWh)', color=THEME['text'], fontsize=16)

    fig.suptitle(f'{payload.title} · Reporte ejecutivo', color=THEME['text'], fontsize=20, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.96])
    fig.savefig(path, dpi=180, bbox_inches='tight', facecolor=fig.get_facecolor())
    plt.close(fig)
    return path
