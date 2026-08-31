from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, time, timedelta
from io import BytesIO
import logging
from typing import Any, Iterable
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError

from app.database import SessionLocal
from app.services.durango_capabilities import (
    DURANGO_SCADA_CUTOVER_LOCAL,
    JARABES,
    JARABES_CHANNEL_CUTOVER_LOCAL,
    JARABES_SOURCE_SEGMENTS,
    LAVADORAS,
    LINE_FLOWS,
    LINES,
    LOCAL_TIMEZONE,
    POZO_1_FLOW_CALIBRATION_CUTOFF_LOCAL,
    WELLS,
    normalize_flow_lps,
)
from app.services.plant_time import local_now_naive, local_to_source_naive, source_to_local_naive
from app.services.water_history_service import WaterHistoryError, get_water_history_module

logger = logging.getLogger(__name__)
LOCAL_ZONE = ZoneInfo(LOCAL_TIMEZONE)

# Confirmado por la auditoría del Incremental 01.
PHYSICAL_HISTORY_START = date(2026, 6, 3)
IOT_PHYSICAL_START_LOCAL = datetime(2026, 6, 3, 15, 35)
VALIDATED_SEGMENT_START_LOCAL = DURANGO_SCADA_CUTOVER_LOCAL
MAX_EXPORT_DAYS = 366


class HistoricalExportError(RuntimeError):
    def __init__(self, message: str, *, status: str = 'sql_error') -> None:
        super().__init__(message)
        self.status = status


def _num(value: Any) -> float | None:
    if value in (None, ''):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if parsed == parsed else None


def _parse_date(value: Any | None, *, default: date) -> date:
    if value in (None, ''):
        return default
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError as exc:
        raise ValueError(f'Fecha no válida: {value}') from exc


def _range(start_date: Any | None = None, end_date: Any | None = None, *, now: datetime | None = None) -> tuple[date, date, datetime]:
    now_local = (now or local_now_naive()).replace(tzinfo=None)
    start = _parse_date(start_date, default=PHYSICAL_HISTORY_START)
    end = _parse_date(end_date, default=now_local.date())
    if start < PHYSICAL_HISTORY_START:
        start = PHYSICAL_HISTORY_START
    if end > now_local.date():
        end = now_local.date()
    if end < start:
        raise ValueError('La fecha final no puede ser anterior a la fecha inicial.')
    if (end - start).days + 1 > MAX_EXPORT_DAYS:
        raise ValueError(f'El histórico completo permite un máximo de {MAX_EXPORT_DAYS} días por exportación.')
    return start, end, now_local


def _date_iter(start: date, end: date):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def _module_label(contract: dict[str, Any]) -> str:
    key = str(contract.get('operational_key') or '')
    if contract in WELLS or str(contract.get('group')) == 'well':
        return 'Pozos'
    if contract in LINES or str(contract.get('group')) == 'line':
        return 'Líneas'
    if key == 'jarabes':
        return 'Jarabes'
    return 'Lavadoras'


def _specs() -> list[dict[str, Any]]:
    specs: list[dict[str, Any]] = []
    for contract in [*WELLS, *LINES, *LINE_FLOWS, *LAVADORAS, *JARABES]:
        source = str(contract.get('table') or 'iot.readings_minute')
        is_iot = contract.get('sensor_id') is not None and source in {'dbo.SensorsBOS_Pozo', 'dbo.SensorsBOS_Linea'}
        specs.append({
            'key': str(contract['operational_key']),
            'name': str(contract['display_name']),
            'module': str(contract.get('group') or 'flow'),
            'module_label': _module_label(contract),
            'sensor_id': int(contract['sensor_id']) if contract.get('sensor_id') is not None else None,
            'source': 'iot.readings_minute' if is_iot else source,
            'source_key': str(contract.get('source_key') or contract.get('operational_key')),
            'source_timezone': str(contract.get('source_timestamp_timezone') or LOCAL_TIMEZONE),
            'raw_flow_unit': str(contract.get('raw_flow_unit') or 'L/s'),
            'flow_encoding': contract.get('flow_encoding'),
            'physical_start_local': IOT_PHYSICAL_START_LOCAL if is_iot else VALIDATED_SEGMENT_START_LOCAL,
            'validated_start_local': VALIDATED_SEGMENT_START_LOCAL,
        })
    return specs


def _effective_day_window(day: date, now_local: datetime) -> tuple[datetime, datetime]:
    start = datetime.combine(day, time.min)
    end = start + timedelta(days=1)
    if day == now_local.date():
        end = min(end, now_local.replace(second=0, microsecond=0) + timedelta(minutes=1))
    return start, end


def _expected_minutes(day: date, spec: dict[str, Any], now_local: datetime) -> int:
    day_start, day_end = _effective_day_window(day, now_local)
    physical_start = spec['physical_start_local']
    effective_start = max(day_start, physical_start)
    if day_end <= effective_start:
        return 0
    return max(1, int((day_end - effective_start).total_seconds() // 60))


def _coverage_pct(records: int, expected: int) -> float:
    if expected <= 0:
        return 0.0
    return round(min(max(records, 0) * 100.0 / expected, 100.0), 2)


def _coverage_status(records: int, expected: int, *, current_day: bool) -> str:
    if expected <= 0:
        return 'Fuera de ventana física'
    if records <= 0:
        return 'Sin registros'
    pct = _coverage_pct(records, expected)
    if current_day and pct >= 99.0:
        return 'Completo hasta el momento'
    if pct >= 99.95:
        return 'Completo'
    if pct >= 95.0:
        return 'Casi completo'
    return 'Cobertura parcial'


def _sql_error(exc: Exception, message: str) -> HistoricalExportError:
    status = 'timeout' if 'timeout' in str(exc).lower() or 'hy008' in str(exc).lower() else 'sql_error'
    return HistoricalExportError(message, status=status)


def _query_iot_minute_rows(session: Any, start_local: datetime, end_local: datetime, sensor_ids: Iterable[int]) -> list[dict[str, Any]]:
    ids = sorted({int(value) for value in sensor_ids})
    if not ids or end_local <= start_local:
        return []
    params: dict[str, Any] = {'start_local': start_local, 'end_local': end_local}
    placeholders: list[str] = []
    for index, sensor_id in enumerate(ids):
        key = f'sensor_{index}'
        params[key] = sensor_id
        placeholders.append(f':{key}')
    sql = text(f'''
        WITH source_rows AS (
            SELECT
                reading.sensor_id,
                COALESCE(reading.ts_local, reading.ts_minute, reading.inserted_at) AS reading_ts,
                TRY_CONVERT(float, reading.instant_value) AS raw_flow,
                TRY_CONVERT(float, reading.total_value) AS total_value,
                reading.quality,
                reading.source,
                reading.inserted_at,
                DATEADD(minute, DATEDIFF(minute, 0, COALESCE(reading.ts_local, reading.ts_minute, reading.inserted_at)), 0) AS minute_ts
            FROM iot.readings_minute AS reading
            WHERE reading.sensor_id IN ({', '.join(placeholders)})
              AND COALESCE(reading.ts_local, reading.ts_minute, reading.inserted_at) >= :start_local
              AND COALESCE(reading.ts_local, reading.ts_minute, reading.inserted_at) < :end_local
              AND (reading.instant_value IS NOT NULL OR reading.total_value IS NOT NULL)
        ), ranked AS (
            SELECT *, ROW_NUMBER() OVER (
                PARTITION BY sensor_id, minute_ts
                ORDER BY reading_ts DESC, inserted_at DESC
            ) AS rn
            FROM source_rows
        )
        SELECT sensor_id, reading_ts, minute_ts, raw_flow, total_value, quality, source
        FROM ranked
        WHERE rn = 1
        ORDER BY minute_ts, sensor_id
    ''')
    try:
        return [dict(row._mapping) for row in session.execute(sql, params).fetchall()]
    except SQLAlchemyError as exc:
        try:
            session.rollback()
        except Exception:
            pass
        raise _sql_error(exc, 'No fue posible consultar el histórico minuto a minuto de Pozos/Líneas.') from exc


def _query_lavadora_minute_rows(session: Any, start_local: datetime, end_local: datetime) -> list[dict[str, Any]]:
    start_local = max(start_local, VALIDATED_SEGMENT_START_LOCAL)
    if end_local <= start_local:
        return []
    start_utc = local_to_source_naive(start_local, 'UTC')
    end_utc = local_to_source_naive(end_local, 'UTC')
    select_columns: list[str] = []
    for contract in LAVADORAS:
        select_columns.extend([
            f"TRY_CONVERT(float, {contract['instant_column']}) AS {contract['operational_key']}_raw_flow",
            f"TRY_CONVERT(float, {contract['total_column']}) AS {contract['operational_key']}_total",
        ])
    sql = text(f'''
        WITH ranked AS (
            SELECT
                Time_Stamp,
                {', '.join(select_columns)},
                ROW_NUMBER() OVER (
                    PARTITION BY DATEADD(minute, DATEDIFF(minute, 0, Time_Stamp), 0)
                    ORDER BY Time_Stamp DESC
                ) AS rn
            FROM dbo.SensorsBOS_Lavadoras
            WHERE Time_Stamp >= :start_utc
              AND Time_Stamp < :end_utc
        )
        SELECT Time_Stamp, {', '.join(column.split(' AS ')[-1] for column in select_columns)}
        FROM ranked
        WHERE rn = 1
        ORDER BY Time_Stamp
    ''')
    try:
        source_rows = [dict(row._mapping) for row in session.execute(sql, {'start_utc': start_utc, 'end_utc': end_utc}).fetchall()]
    except SQLAlchemyError as exc:
        try:
            session.rollback()
        except Exception:
            pass
        raise _sql_error(exc, 'No fue posible consultar el histórico minuto a minuto de Lavadoras.') from exc

    output: list[dict[str, Any]] = []
    for row in source_rows:
        local_stamp = source_to_local_naive(row.get('Time_Stamp'), 'UTC')
        if local_stamp is None:
            continue
        minute = local_stamp.replace(second=0, microsecond=0)
        for contract in LAVADORAS:
            key = str(contract['operational_key'])
            raw_flow = _num(row.get(f'{key}_raw_flow'))
            total = _num(row.get(f'{key}_total'))
            if raw_flow is None and total is None:
                continue
            output.append({
                'operational_key': key,
                'reading_ts': local_stamp,
                'minute_ts': minute,
                'raw_flow': raw_flow,
                'flow_lps': normalize_flow_lps(key, raw_flow, local_stamp),
                'total_value': total,
                'quality': None,
                'source': 'dbo.SensorsBOS_Lavadoras',
                'source_key': str(contract.get('source_key') or key),
            })
    return output


def _query_jarabes_minute_rows(session: Any, start_local: datetime, end_local: datetime) -> list[dict[str, Any]]:
    start_local = max(start_local, VALIDATED_SEGMENT_START_LOCAL)
    if end_local <= start_local:
        return []
    start_utc = local_to_source_naive(start_local, 'UTC')
    end_utc = local_to_source_naive(end_local, 'UTC')
    output: list[dict[str, Any]] = []
    for segment in JARABES_SOURCE_SEGMENTS:
        segment_start = max(start_utc, segment.get('start_utc') or datetime.min)
        segment_end = min(end_utc, segment.get('end_utc') or datetime.max)
        if segment_end <= segment_start:
            continue
        instant_column = str(segment['instant_column'])
        total_column = str(segment['total_column'])
        sensor_column = str(segment['sensor_column'])
        sql = text(f'''
            WITH ranked AS (
                SELECT
                    Time_Stamp,
                    TRY_CONVERT(float, {sensor_column}) AS source_sensor_id,
                    TRY_CONVERT(float, {instant_column}) AS raw_flow,
                    TRY_CONVERT(float, {total_column}) AS total_value,
                    ROW_NUMBER() OVER (
                        PARTITION BY DATEADD(minute, DATEDIFF(minute, 0, Time_Stamp), 0)
                        ORDER BY Time_Stamp DESC
                    ) AS rn
                FROM dbo.SensorsBOS_Tanque
                WHERE Time_Stamp >= :start_utc
                  AND Time_Stamp < :end_utc
                  AND (TRY_CONVERT(int, {sensor_column}) = :sensor_id OR TRY_CONVERT(int, {sensor_column}) IS NULL)
            )
            SELECT Time_Stamp, source_sensor_id, raw_flow, total_value
            FROM ranked
            WHERE rn = 1
            ORDER BY Time_Stamp
        ''')
        try:
            rows = session.execute(sql, {
                'start_utc': segment_start,
                'end_utc': segment_end,
                'sensor_id': int(segment['sensor_id']),
            }).fetchall()
        except SQLAlchemyError as exc:
            try:
                session.rollback()
            except Exception:
                pass
            raise _sql_error(exc, 'No fue posible consultar el histórico minuto a minuto de Jarabes.') from exc
        for source_row in rows:
            row = dict(source_row._mapping)
            local_stamp = source_to_local_naive(row.get('Time_Stamp'), 'UTC')
            if local_stamp is None:
                continue
            raw_flow = _num(row.get('raw_flow'))
            output.append({
                'operational_key': 'jarabes',
                'reading_ts': local_stamp,
                'minute_ts': local_stamp.replace(second=0, microsecond=0),
                'raw_flow': raw_flow,
                'flow_lps': normalize_flow_lps('jarabes', raw_flow, local_stamp),
                'total_value': _num(row.get('total_value')),
                'quality': None,
                'source': 'dbo.SensorsBOS_Tanque',
                'source_key': str(segment['source_key']),
                'source_sensor_id': int(segment['sensor_id']),
            })
    output.sort(key=lambda row: row['minute_ts'])
    return output


def _query_day_minutes(session: Any, day_start: datetime, day_end: datetime, specs: list[dict[str, Any]]) -> dict[str, dict[datetime, dict[str, Any]]]:
    grouped: dict[str, dict[datetime, dict[str, Any]]] = {str(spec['key']): {} for spec in specs}
    iot_specs = [spec for spec in specs if spec['source'] == 'iot.readings_minute' and spec.get('sensor_id') is not None]
    id_to_spec = {int(spec['sensor_id']): spec for spec in iot_specs}
    for row in _query_iot_minute_rows(session, day_start, day_end, id_to_spec.keys()):
        sensor_id = int(row['sensor_id'])
        spec = id_to_spec.get(sensor_id)
        if spec is None:
            continue
        minute = row['minute_ts'].replace(second=0, microsecond=0)
        raw_flow = _num(row.get('raw_flow'))
        grouped[str(spec['key'])][minute] = {
            'operational_key': spec['key'],
            'reading_ts': row.get('reading_ts'),
            'minute_ts': minute,
            'raw_flow': raw_flow,
            'flow_lps': normalize_flow_lps(sensor_id, raw_flow, minute),
            'total_value': _num(row.get('total_value')),
            'quality': row.get('quality'),
            'source': row.get('source') or 'iot.readings_minute',
            'source_key': spec['source_key'],
        }
    for row in _query_lavadora_minute_rows(session, day_start, day_end):
        grouped[str(row['operational_key'])][row['minute_ts']] = row
    for row in _query_jarabes_minute_rows(session, day_start, day_end):
        grouped['jarabes'][row['minute_ts']] = row
    return grouped


def _raw_daily_row(spec: dict[str, Any], day: date, rows: dict[datetime, dict[str, Any]], expected: int, now_local: datetime) -> dict[str, Any]:
    ordered = [rows[key] for key in sorted(rows)]
    raw_flows = [_num(row.get('raw_flow')) for row in ordered]
    raw_flows = [value for value in raw_flows if value is not None]
    normalized = [_num(row.get('flow_lps')) for row in ordered]
    normalized = [value for value in normalized if value is not None]
    totals = [_num(row.get('total_value')) for row in ordered]
    totals = [value for value in totals if value is not None]
    opening = totals[0] if totals else None
    closing = totals[-1] if totals else None
    return {
        'date': day,
        'module_label': spec['module_label'],
        'name': spec['name'],
        'operational_key': spec['key'],
        'sensor_id': spec['sensor_id'],
        'raw_flow_avg': (sum(raw_flows) / len(raw_flows)) if raw_flows else None,
        'flow_avg_lps': (sum(normalized) / len(normalized)) if normalized else None,
        'flow_min_lps': min(normalized) if normalized else None,
        'flow_max_lps': max(normalized) if normalized else None,
        'opening_m3': opening,
        'closing_m3': closing,
        'raw_delta_m3': (closing - opening) if opening is not None and closing is not None else None,
        'samples': len(ordered),
        'expected_samples': expected,
        'coverage_pct': _coverage_pct(len(ordered), expected),
        'coverage_status': _coverage_status(len(ordered), expected, current_day=(day == now_local.date())),
        'source': spec['source'],
    }


def _raw_summary_update(summary: dict[str, Any], row: dict[str, Any], minute_rows: dict[datetime, dict[str, Any]]) -> None:
    summary['samples'] += row['samples']
    summary['expected_samples'] += row['expected_samples']
    ordered = [minute_rows[key] for key in sorted(minute_rows)]
    if not ordered:
        return
    first = ordered[0]
    last = ordered[-1]
    if summary['first_timestamp'] is None:
        summary['first_timestamp'] = first['minute_ts']
        first_raw = next((_num(item.get('raw_flow')) for item in ordered if _num(item.get('raw_flow')) is not None), None)
        first_flow = next((_num(item.get('flow_lps')) for item in ordered if _num(item.get('flow_lps')) is not None), None)
        summary['first_raw_flow'] = first_raw
        summary['first_flow_lps'] = first_flow
    summary['last_timestamp'] = last['minute_ts']
    last_raw = next((_num(item.get('raw_flow')) for item in reversed(ordered) if _num(item.get('raw_flow')) is not None), None)
    last_flow = next((_num(item.get('flow_lps')) for item in reversed(ordered) if _num(item.get('flow_lps')) is not None), None)
    if last_raw is not None:
        summary['last_raw_flow'] = last_raw
    if last_flow is not None:
        summary['last_flow_lps'] = last_flow
    if summary['opening_m3'] is None and row.get('opening_m3') is not None:
        summary['opening_m3'] = row['opening_m3']
    if row.get('closing_m3') is not None:
        summary['closing_m3'] = row['closing_m3']


def _raw_period_summary(specs: list[dict[str, Any]], raw_daily: list[dict[str, Any]], summaries: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for spec in specs:
        state = summaries[str(spec['key'])]
        opening = state['opening_m3']
        closing = state['closing_m3']
        expected = int(state['expected_samples'])
        samples = int(state['samples'])
        output.append({
            **spec,
            **state,
            'raw_delta_m3': (closing - opening) if opening is not None and closing is not None else None,
            'coverage_pct': _coverage_pct(samples, expected),
        })
    return output


def _hydraulic_daily_rows(start: date, end: date) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for module in ('well', 'line', 'flow'):
        try:
            payload = get_water_history_module(
                module=module,
                start_date=start.isoformat(),
                end_date=end.isoformat(),
                aggregation='daily',
                force_refresh=False,
            )
        except (WaterHistoryError, ValueError) as exc:
            logger.warning('Durango full historical export daily reconciliation failed module=%s error=%s', module, exc)
            continue
        for series in payload.get('series') or []:
            op_key = str(series.get('operational_key') or series.get('sensor_id') or '')
            for point in series.get('points') or []:
                bucket = str(point.get('bucket_start') or '')
                try:
                    day = date.fromisoformat(bucket[:10])
                except ValueError:
                    continue
                output.append({
                    'date': day,
                    'module': module,
                    'name': series.get('name'),
                    'operational_key': op_key,
                    'sensor_id': series.get('sensor_id'),
                    'opening_m3': point.get('totalizer_open_m3'),
                    'closing_m3': point.get('totalizer_close_m3'),
                    'validated_volume_m3': point.get('validated_volume_m3'),
                    'volume_m3': point.get('volume_m3'),
                    'volume_reliable': bool(point.get('volume_reliable')),
                    'flow_avg_lps': point.get('flow_avg_lps'),
                    'samples_received': point.get('samples_received'),
                    'samples_expected': point.get('samples_expected'),
                    'coverage_percent': point.get('coverage_percent'),
                    'coverage_status': point.get('coverage_status'),
                    'data_status': point.get('data_status'),
                    'validation_status': point.get('validation_status'),
                    'discarded_events': point.get('discarded_totalizer_events') or 0,
                })
    output.sort(key=lambda row: (row['date'], row['module'], str(row['name'])))
    return output


def _hydraulic_period_summary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = str(row['operational_key'])
        item = grouped.setdefault(key, {
            'name': row['name'], 'module': row['module'], 'sensor_id': row['sensor_id'],
            'validated_volume_m3': 0.0, 'validated_days': 0, 'review_days': 0, 'no_data_days': 0,
        })
        if row['volume_reliable'] and row['validated_volume_m3'] is not None:
            item['validated_volume_m3'] += float(row['validated_volume_m3'])
            item['validated_days'] += 1
        elif str(row.get('data_status') or '').lower() in {'no_data', 'legacy_configuration_pending'}:
            item['no_data_days'] += 1
        else:
            item['review_days'] += 1
    return list(grouped.values())


def _header(ws: Any, values: list[str], fill: str = '0B4F79') -> None:
    cells = []
    for value in values:
        cell = WriteOnlyCell(ws, value=value)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cells.append(cell)
    ws.append(cells)


def _add_summary_sheet(wb: Workbook, start: date, end: date, now_local: datetime, specs: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet('Resumen')
    _header(ws, ['Campo', 'Valor'])
    rows = [
        ('Planta', 'Durango'),
        ('Periodo exportado', f'{start.isoformat()} a {end.isoformat()}'),
        ('Generado', now_local.isoformat(sep=' ', timespec='minutes')),
        ('Inicio físico confirmado iot.readings_minute', '03/06/2026 15:35'),
        ('Inicio segmento hidráulico validado', VALIDATED_SEGMENT_START_LOCAL.isoformat(sep=' ', timespec='minutes')),
        ('Pozo 1 cambio de unidad', POZO_1_FLOW_CALIBRATION_CUTOFF_LOCAL.isoformat(sep=' ', timespec='minutes')),
        ('Jarabes cambio de canal', JARABES_CHANNEL_CUTOVER_LOCAL.isoformat(sep=' ', timespec='minutes')),
        ('Elementos', len(specs)),
        ('Nota', 'Las hojas crudas conservan los valores observados. Las hojas conciliadas aplican el contrato hidráulico y de calidad de Durango.'),
    ]
    for row in rows:
        ws.append(list(row))
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 92


def _add_raw_summary_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet('Resumen crudo')
    _header(ws, ['Módulo', 'Elemento', 'Clave', 'Sensor', 'Primera muestra', 'Última muestra', 'Flujo bruto inicial', 'Flujo bruto final', 'Flujo L/s inicial', 'Flujo L/s final', 'Totalizador inicial m³', 'Totalizador final m³', 'Delta crudo m³', 'Muestras', 'Esperadas', 'Cobertura %'])
    for row in rows:
        ws.append([
            row['module_label'], row['name'], row['key'], row['sensor_id'], row['first_timestamp'], row['last_timestamp'],
            row['first_raw_flow'], row['last_raw_flow'], row['first_flow_lps'], row['last_flow_lps'],
            row['opening_m3'], row['closing_m3'], row['raw_delta_m3'], row['samples'], row['expected_samples'], row['coverage_pct'],
        ])


def _add_raw_daily_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet('Datos crudos diarios')
    _header(ws, ['Fecha', 'Módulo', 'Elemento', 'Clave', 'Sensor', 'Flujo bruto promedio', 'Flujo normalizado L/s', 'Flujo mínimo L/s', 'Flujo máximo L/s', 'Total inicial m³', 'Total final m³', 'Delta crudo m³', 'Muestras', 'Esperadas', 'Cobertura %', 'Estado cobertura', 'Fuente'])
    for row in rows:
        ws.append([
            row['date'], row['module_label'], row['name'], row['operational_key'], row['sensor_id'], row['raw_flow_avg'],
            row['flow_avg_lps'], row['flow_min_lps'], row['flow_max_lps'], row['opening_m3'], row['closing_m3'], row['raw_delta_m3'],
            row['samples'], row['expected_samples'], row['coverage_pct'], row['coverage_status'], row['source'],
        ])


def _add_hydraulic_sheets(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    summary_ws = wb.create_sheet('Resumen conciliado')
    _header(summary_ws, ['Módulo', 'Elemento', 'Sensor', 'Volumen validado m³', 'Días validados', 'Días en revisión', 'Días sin datos'])
    module_labels = {'well': 'Pozos', 'line': 'Líneas', 'flow': 'Lavadoras / Jarabes'}
    for row in _hydraulic_period_summary(rows):
        summary_ws.append([
            module_labels.get(str(row['module']), row['module']), row['name'], row['sensor_id'], round(float(row['validated_volume_m3']), 6),
            row['validated_days'], row['review_days'], row['no_data_days'],
        ])

    ws = wb.create_sheet('Conciliado diario')
    _header(ws, ['Fecha', 'Módulo', 'Elemento', 'Clave', 'Sensor', 'Apertura m³', 'Cierre m³', 'Volumen calculado m³', 'Volumen validado m³', 'Confiable', 'Flujo prom L/s', 'Muestras', 'Esperadas', 'Cobertura %', 'Estado cobertura', 'Estado dato', 'Validación', 'Eventos descartados'])
    for row in rows:
        ws.append([
            row['date'], module_labels.get(str(row['module']), row['module']), row['name'], row['operational_key'], row['sensor_id'],
            row['opening_m3'], row['closing_m3'], row['volume_m3'], row['validated_volume_m3'], 'Sí' if row['volume_reliable'] else 'No',
            row['flow_avg_lps'], row['samples_received'], row['samples_expected'], row['coverage_percent'], row['coverage_status'],
            row['data_status'], row['validation_status'], row['discarded_events'],
        ])


def _add_coverage_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet('Cobertura diaria')
    _header(ws, ['Fecha', 'Módulo', 'Elemento', 'Clave', 'Muestras por minuto', 'Esperadas', 'Cobertura %', 'Estado'])
    for row in rows:
        ws.append([row['date'], row['module_label'], row['name'], row['operational_key'], row['samples'], row['expected_samples'], row['coverage_pct'], row['coverage_status']])


def _add_gaps_sheet(wb: Workbook, gap_rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet('Huecos')
    _header(ws, ['Fecha', 'Cobertura mínima %', 'Cobertura máxima %', 'Elementos afectados', 'Tipo', 'Nota'])
    for row in gap_rows:
        ws.append([row['date'], row['min_pct'], row['max_pct'], row['affected'], row['type'], row['note']])


def _add_sensors_sheet(wb: Workbook, specs: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet('Sensores')
    _header(ws, ['Módulo', 'Elemento', 'Clave', 'Sensor', 'Fuente', 'Canal', 'Reloj fuente', 'Unidad flujo bruto', 'Inicio físico usado', 'Inicio validado', 'Regla'])
    for spec in specs:
        if spec['key'] == 'pozo_1':
            rule = f"Antes de {POZO_1_FLOW_CALIBRATION_CUTOFF_LOCAL.isoformat(sep=' ', timespec='minutes')}: flujo bruto m³/h / 3.6; después L/s directo."
        elif spec['key'] == 'jarabes':
            rule = f"Identidad única Jarabes; cambia de canal el {JARABES_CHANNEL_CUTOVER_LOCAL.isoformat(sep=' ', timespec='seconds')}. El flujo bruto codificado se conserva y se muestra también normalizado."
        elif spec['source'] != 'iot.readings_minute':
            rule = 'Timestamp UTC normalizado a America/Mexico_City. Mapeo operativo validado desde el corte SCADA.'
        else:
            rule = 'Timestamp local. Datos anteriores al 04/08 se conservan sólo como crudos; no se relabelan como segmento validado.'
        ws.append([
            spec['module_label'], spec['name'], spec['key'], spec['sensor_id'], spec['source'], spec['source_key'], spec['source_timezone'], spec['raw_flow_unit'],
            spec['physical_start_local'], spec['validated_start_local'], rule,
        ])


def _add_notes_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet('Notas')
    _header(ws, ['Regla', 'Detalle'])
    notes = [
        ('Crudo vs conciliado', 'Crudo conserva lo observado por minuto. Conciliado aplica fronteras [T0,T1), totalizador, cobertura y contrato de calidad.'),
        ('Ceros', 'Un 0 almacenado se conserva como 0. Una ausencia de registro permanece vacía; nunca se reemplaza por 0.'),
        ('Pre-corte', 'Pozos/Líneas/Lavadora Línea 2 pueden tener registros desde 03/06/2026. Antes del 04/08/2026 18:16 se consideran crudos históricos, no segmento operativo validado.'),
        ('Lavadoras/Jarabes', 'El mapeo BOS se exporta desde el corte general validado. No se atribuye a la identidad actual información anterior sin una auditoría física específica.'),
        ('Pozo 1', 'Se conserva el valor bruto y, en paralelo, el flujo normalizado según el corte de calibración del 11/08/2026 12:15.'),
        ('Jarabes', 'Se conserva una sola identidad operativa y se resuelve el canal 3010→3004 por fecha. El flujo almacenado codificado se conserva en columna cruda y se decodifica en la columna normalizada.'),
        ('Cobertura', 'El día actual se evalúa contra los minutos transcurridos. El primer día se evalúa desde el primer instante físico aplicable a cada elemento.'),
    ]
    for row in notes:
        ws.append(list(row))


def build_full_historical_excel(start_date: Any | None = None, end_date: Any | None = None, *, now: datetime | None = None) -> tuple[bytes, str]:
    start, end, now_local = _range(start_date, end_date, now=now)
    specs = _specs()
    spec_by_key = {str(spec['key']): spec for spec in specs}

    wb = Workbook(write_only=True)
    _add_summary_sheet(wb, start, end, now_local, specs)

    raw_wide = wb.create_sheet('Crudo 1 min')
    wide_headers = ['Fecha hora local']
    for spec in specs:
        wide_headers.extend([
            f"{spec['name']} | flujo bruto",
            f"{spec['name']} | flujo L/s",
            f"{spec['name']} | total m³",
            f"{spec['name']} | fuente",
        ])
    _header(raw_wide, wide_headers)

    raw_daily: list[dict[str, Any]] = []
    coverage_rows: list[dict[str, Any]] = []
    gap_rows: list[dict[str, Any]] = []
    summaries: dict[str, dict[str, Any]] = {
        str(spec['key']): {
            'first_timestamp': None, 'last_timestamp': None, 'first_raw_flow': None, 'last_raw_flow': None,
            'first_flow_lps': None, 'last_flow_lps': None, 'opening_m3': None, 'closing_m3': None,
            'samples': 0, 'expected_samples': 0,
        }
        for spec in specs
    }

    session = SessionLocal()
    try:
        for day in _date_iter(start, end):
            day_start, day_end = _effective_day_window(day, now_local)
            if day_end <= day_start:
                continue
            grouped = _query_day_minutes(session, day_start, day_end, specs)
            day_coverages: list[float] = []
            affected = 0
            eligible = 0
            for spec in specs:
                key = str(spec['key'])
                expected = _expected_minutes(day, spec, now_local)
                rows = grouped.get(key, {})
                raw_row = _raw_daily_row(spec, day, rows, expected, now_local)
                raw_daily.append(raw_row)
                _raw_summary_update(summaries[key], raw_row, rows)
                coverage_rows.append(raw_row)
                if expected > 0:
                    eligible += 1
                    day_coverages.append(raw_row['coverage_pct'])
                    if raw_row['coverage_status'] not in {'Completo', 'Completo hasta el momento'}:
                        affected += 1

            if eligible and affected:
                min_pct = min(day_coverages) if day_coverages else 0.0
                max_pct = max(day_coverages) if day_coverages else 0.0
                if max_pct <= 0:
                    gap_type = 'Sin registros general'
                    note = 'No existen registros en los elementos cuya ventana física ya estaba activa.'
                elif day == now_local.date():
                    gap_type = 'Día actual'
                    note = 'Cobertura calculada contra los minutos transcurridos hasta la generación.'
                elif min_pct >= 95.0:
                    gap_type = 'Casi completo'
                    note = 'Existen faltantes menores; conservar como cobertura no íntegra.'
                else:
                    gap_type = 'Cobertura parcial'
                    note = 'Existe una interrupción de adquisición/almacenamiento en uno o más elementos.'
                gap_rows.append({'date': day, 'min_pct': min_pct, 'max_pct': max_pct, 'affected': affected, 'type': gap_type, 'note': note})

            cursor = max(day_start, min(spec['physical_start_local'] for spec in specs))
            while cursor < day_end:
                row_values: list[Any] = [cursor]
                for spec in specs:
                    sample = grouped.get(str(spec['key']), {}).get(cursor)
                    row_values.extend([
                        sample.get('raw_flow') if sample else None,
                        sample.get('flow_lps') if sample else None,
                        sample.get('total_value') if sample else None,
                        sample.get('source_key') if sample else None,
                    ])
                raw_wide.append(row_values)
                cursor += timedelta(minutes=1)
    finally:
        session.close()

    raw_period = _raw_period_summary(specs, raw_daily, summaries)
    _add_raw_summary_sheet(wb, raw_period)
    _add_raw_daily_sheet(wb, raw_daily)
    hydraulic_rows = _hydraulic_daily_rows(start, end)
    _add_hydraulic_sheets(wb, hydraulic_rows)
    _add_coverage_sheet(wb, coverage_rows)
    _add_gaps_sheet(wb, gap_rows)
    _add_sensors_sheet(wb, specs)
    _add_notes_sheet(wb)

    buffer = BytesIO()
    wb.save(buffer)
    filename = f'Historico_Completo_Durango_{start.isoformat()}_{end.isoformat()}.xlsx'
    return buffer.getvalue(), filename


def _compress_ranges(days: list[date]) -> list[str]:
    if not days:
        return []
    ordered = sorted(set(days))
    output: list[str] = []
    start = previous = ordered[0]
    for current in ordered[1:]:
        if current == previous + timedelta(days=1):
            previous = current
            continue
        output.append(start.strftime('%d/%m/%Y') if start == previous else f'{start.strftime("%d/%m/%Y")}–{previous.strftime("%d/%m/%Y")}')
        start = previous = current
    output.append(start.strftime('%d/%m/%Y') if start == previous else f'{start.strftime("%d/%m/%Y")}–{previous.strftime("%d/%m/%Y")}')
    return output


def _coverage_and_raw_summary(start: date, end: date, now_local: datetime, specs: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    coverage_rows: list[dict[str, Any]] = []
    raw_daily: list[dict[str, Any]] = []
    summaries: dict[str, dict[str, Any]] = {
        str(spec['key']): {
            'first_timestamp': None, 'last_timestamp': None, 'first_raw_flow': None, 'last_raw_flow': None,
            'first_flow_lps': None, 'last_flow_lps': None, 'opening_m3': None, 'closing_m3': None,
            'samples': 0, 'expected_samples': 0,
        }
        for spec in specs
    }
    session = SessionLocal()
    try:
        for day in _date_iter(start, end):
            day_start, day_end = _effective_day_window(day, now_local)
            grouped = _query_day_minutes(session, day_start, day_end, specs)
            for spec in specs:
                key = str(spec['key'])
                expected = _expected_minutes(day, spec, now_local)
                row = _raw_daily_row(spec, day, grouped.get(key, {}), expected, now_local)
                raw_daily.append(row)
                coverage_rows.append(row)
                _raw_summary_update(summaries[key], row, grouped.get(key, {}))
    finally:
        session.close()
    return coverage_rows, raw_daily, _raw_period_summary(specs, raw_daily, summaries)


def build_full_historical_pdf(start_date: Any | None = None, end_date: Any | None = None, *, now: datetime | None = None) -> tuple[bytes, str]:
    start, end, now_local = _range(start_date, end_date, now=now)
    specs = _specs()
    coverage_rows, raw_daily, raw_summary = _coverage_and_raw_summary(start, end, now_local, specs)
    hydraulic_rows = _hydraulic_daily_rows(start, end)
    hydraulic_summary = _hydraulic_period_summary(hydraulic_rows)

    day_groups: dict[date, list[dict[str, Any]]] = defaultdict(list)
    for row in coverage_rows:
        if row['expected_samples'] > 0:
            day_groups[row['date']].append(row)
    zero_days: list[date] = []
    partial_days: list[date] = []
    daily_global: list[dict[str, Any]] = []
    for day in _date_iter(start, end):
        rows = day_groups.get(day, [])
        if not rows:
            continue
        min_pct = min(float(row['coverage_pct']) for row in rows)
        max_pct = max(float(row['coverage_pct']) for row in rows)
        missing = sum(1 for row in rows if row['samples'] == 0)
        affected = sum(1 for row in rows if row['coverage_status'] not in {'Completo', 'Completo hasta el momento'})
        if missing == len(rows):
            zero_days.append(day)
            status = 'Sin registros general'
        elif affected:
            partial_days.append(day)
            status = 'Día actual' if day == now_local.date() else ('Casi completo' if min_pct >= 95 else 'Cobertura parcial')
        else:
            status = 'Completo hasta el momento' if day == now_local.date() else 'Completo'
        daily_global.append({'date': day, 'min_pct': min_pct, 'max_pct': max_pct, 'affected': affected, 'status': status})

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
        title='Histórico completo de agua - Planta Durango', author='Dashboard ARCA',
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle('TitleDurangoHistory', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=20, leading=24, textColor=colors.HexColor('#0B3558'))
    h2 = ParagraphStyle('H2DurangoHistory', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, leading=15, textColor=colors.HexColor('#0B4F79'), spaceBefore=6, spaceAfter=5)
    body = ParagraphStyle('BodyDurangoHistory', parent=styles['BodyText'], fontSize=8.3, leading=10.5, textColor=colors.HexColor('#20364A'))

    story: list[Any] = [
        Paragraph('Histórico completo de agua · Planta Durango', title),
        Paragraph(f'Periodo: {start.strftime("%d/%m/%Y")} a {end.strftime("%d/%m/%Y")} · Generado: {now_local.strftime("%d/%m/%Y %H:%M")}', body),
        Spacer(1, 4 * mm),
        Paragraph('Criterios de interpretación', h2),
        Paragraph(
            'El histórico físico confirmado de Pozos/Líneas inicia el 03/06/2026 15:35. El segmento hidráulico homologado y validado inicia el 04/08/2026 18:16. '
            'Los datos anteriores se conservan como evidencia cruda y no se relabelan como parte del segmento validado. Lavadoras y Jarabes se atribuyen a su identidad operativa actual únicamente desde el corte general validado.',
            body,
        ),
        Paragraph(
            'Un cero almacenado se conserva como 0; un hueco permanece vacío. El Excel incluye valores crudos por minuto y datos conciliados diarios por separado.',
            body,
        ),
        Spacer(1, 3 * mm),
        Paragraph('Cortes conocidos', h2),
    ]
    cutovers = [
        ['Corte', 'Fecha local', 'Regla'],
        ['SCADA general', '04/08/2026 18:16', 'Inicio del segmento operacional validado'],
        ['Pozo 1', '11/08/2026 12:15', 'Antes: flujo bruto m³/h / 3.6; después: L/s directo'],
        ['Jarabes', '11/08/2026 13:40:29', 'TANQUE_FLOW_IN[4] / 3010 → TANQUE_FLOW_IN[1] / 3004'],
    ]
    table = Table(cutovers, colWidths=[40*mm, 42*mm, 120*mm], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0B4F79')), ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'), ('FONTSIZE',(0,0),(-1,-1),7),
        ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#B8D5E6')), ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F4F9FC')]),
        ('TEXTCOLOR',(0,1),(-1,-1),colors.HexColor('#20364A')), ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]))
    story.extend([table, Spacer(1, 4 * mm), Paragraph('Resumen crudo del periodo', h2)])

    raw_table = [['Módulo','Elemento','Primera','Última','Total inicial','Total final','Delta crudo','Muestras','Cobertura']]
    for row in raw_summary:
        raw_table.append([
            row['module_label'], row['name'],
            '—' if row['first_timestamp'] is None else row['first_timestamp'].strftime('%d/%m %H:%M'),
            '—' if row['last_timestamp'] is None else row['last_timestamp'].strftime('%d/%m %H:%M'),
            '—' if row['opening_m3'] is None else f"{float(row['opening_m3']):,.2f}",
            '—' if row['closing_m3'] is None else f"{float(row['closing_m3']):,.2f}",
            '—' if row['raw_delta_m3'] is None else f"{float(row['raw_delta_m3']):,.2f}",
            str(row['samples']), f"{float(row['coverage_pct']):.2f}%",
        ])
    raw_tbl = Table(raw_table, colWidths=[24*mm,34*mm,27*mm,27*mm,30*mm,30*mm,27*mm,22*mm,25*mm], repeatRows=1)
    raw_tbl.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#0B4F79')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),6.3),('GRID',(0,0),(-1,-1),0.2,colors.HexColor('#C8DCE8')),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F4F9FC')]),
        ('TEXTCOLOR',(0,1),(-1,-1),colors.HexColor('#20364A')),('VALIGN',(0,0),(-1,-1),'MIDDLE')
    ]))
    story.append(raw_tbl)

    story.extend([Spacer(1, 4 * mm), Paragraph('Resumen conciliado', h2)])
    hydraulic_table = [['Módulo','Elemento','Volumen validado','Días validados','En revisión','Sin datos']]
    module_labels = {'well': 'Pozos', 'line': 'Líneas', 'flow': 'Lavadoras / Jarabes'}
    for row in hydraulic_summary:
        hydraulic_table.append([
            module_labels.get(str(row['module']), row['module']), row['name'], f"{float(row['validated_volume_m3']):,.2f} m³",
            row['validated_days'], row['review_days'], row['no_data_days'],
        ])
    hyd_tbl = Table(hydraulic_table, colWidths=[36*mm,48*mm,38*mm,32*mm,32*mm,32*mm], repeatRows=1)
    hyd_tbl.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#0B4F79')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),6.8),('GRID',(0,0),(-1,-1),0.2,colors.HexColor('#C8DCE8')),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F4F9FC')]),
        ('TEXTCOLOR',(0,1),(-1,-1),colors.HexColor('#20364A')),('VALIGN',(0,0),(-1,-1),'MIDDLE')
    ]))
    story.append(hyd_tbl)

    story.extend([PageBreak(), Paragraph('Cobertura diaria', h2)])
    zero_ranges = _compress_ranges(zero_days)
    if zero_ranges:
        story.append(Paragraph('Días sin registros en todos los elementos cuya ventana física estaba activa: ' + ', '.join(zero_ranges) + '.', body))
    else:
        story.append(Paragraph('No se detectaron días completos sin registros para todos los elementos activos dentro del periodo.', body))
    if partial_days:
        story.append(Paragraph(f'Días con cobertura no íntegra: {len(set(partial_days))}. No deben interpretarse como días completos.', body))
    coverage_table = [['Fecha','Cob. mínima','Cob. máxima','Elementos afectados','Estado']]
    for row in daily_global:
        coverage_table.append([row['date'].strftime('%d/%m/%Y'), f"{row['min_pct']:.2f}%", f"{row['max_pct']:.2f}%", row['affected'], row['status']])
    cov_tbl = Table(coverage_table, colWidths=[30*mm,30*mm,30*mm,38*mm,72*mm], repeatRows=1)
    cov_tbl.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#0B4F79')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),7),('GRID',(0,0),(-1,-1),0.2,colors.HexColor('#C8DCE8')),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F4F9FC')]),
        ('TEXTCOLOR',(0,1),(-1,-1),colors.HexColor('#20364A')),('VALIGN',(0,0),(-1,-1),'MIDDLE')
    ]))
    story.append(cov_tbl)

    story.extend([PageBreak(), Paragraph('Notas', h2)])
    notes = [
        'El PDF resume el histórico; el Excel contiene el detalle crudo minuto a minuto y las hojas conciliadas.',
        'Crudo no significa confiable para balance: conserva lo almacenado incluso cuando el contrato hidráulico marque cobertura parcial o revisión.',
        'Los totalizadores crudos pueden contener resets o saltos; el delta crudo no sustituye al volumen conciliado.',
        'El día actual se compara contra los minutos transcurridos y no contra 1440 minutos.',
        'Pozo 1 conserva flujo bruto y flujo normalizado para documentar el cambio de calibración.',
        'Jarabes conserva una identidad lógica única y resuelve el canal físico según la fecha de corte.',
    ]
    for item in notes:
        story.append(Paragraph('• ' + item, body))

    doc.build(story)
    filename = f'Historico_Completo_Durango_{start.isoformat()}_{end.isoformat()}.pdf'
    return buffer.getvalue(), filename
